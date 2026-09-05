"""Dual-track calligraphy label pilot (Gemini 3.6 + Paddle v6) → Task.xlsx B2.

Pilot gán nhãn thư pháp hai nhánh; output Task.xlsx B2.
Writes only under ocr/label_dual_pilot/ — never ocr_result.jsonl /
Chỉ ghi ocr/label_dual_pilot/ — không đụng ocr_result.jsonl.
"""
from __future__ import annotations

import base64
import json
import os
import re
import shutil
import tempfile
import threading
import time
import statistics
from collections import Counter
from collections.abc import Iterator
from concurrent.futures import ThreadPoolExecutor, Future
from difflib import SequenceMatcher
from io import BytesIO
from typing import Any, Callable

from common.api_keys import collect_api_keys
from common.chau_ban_schema import extract_json_object, utc_now_iso
from common.config import get_value, load_config
from common.io_storage import (
    get_minio_client,
    list_objects_with_prefix,
    object_exists,
    upload_file,
    upload_json_payload,
)
from common.llm_chat import call_chat_completion

from final_exam_nlp_crawl_runner import _group_root, _read_json_object
from final_exam_nlp_ocr import (
    _object_key_from_image_path,
    _ocr_output_key,
    _read_object_bytes,
    _to_png_bytes,
    _vision_ocr,
    preprocess_light,
)
from final_exam_nlp_ocr_eval import compact_cjk
from fen_crawl_common import (
    ensure_raw_bucket,
    permalink_for,
    read_jsonl,
    task_export_valid_key,
    write_jsonl,
)

LOG = "[fen_label_dual]"
SCHEMA = "label-dual-1.6.0"
GLM_LOG_SCHEMA = "label-dual-glm-log-1.3"
GLM_MODEL = "glm-5.3-flash"
# Page flags that duplicate status or old noise / Flag trùng status hoặc nhiễu cũ
DROP_PAGE_FLAGS = frozenset({"needs_hitl", "eval_fail"})
# Any of these blocks silver / Cờ nào trong này cũng chặn silver
BLOCK_SILVER_FLAGS = frozenset(
    {
        "weak_evidence",
        "ui_chrome",
        "printed_script",
        "latin_heavy",
        "per_char_boxes",
        "ocr_suspect",
        "low_box_conf",
        "screenshot",
        "not_calligraphy",
        "ocr_unreadable",
        "seal_only",
        "gemini_empty_ink",
    }
)
CJK_RE = re.compile(r"[\u3400-\u4DBF\u4E00-\u9FFF\uF900-\uFAFF]")
LATIN_RE = re.compile(r"[A-Za-z]")
UI_RE = re.compile(
    r"https?://|baijiahao|facebook\.com|\bLTE\b|android|UI\s*Chrome|UI\s*Navigation",
    re.I,
)
_CAPTION_URL_RE = re.compile(r"https?://\S+", re.I)
_CAPTION_HASH_RE = re.compile(r"[#＃][\w\u3400-\u9fff]+")
_CAPTION_MENTION_RE = re.compile(r"@[\w.]+")
_THINK_RE = re.compile(r"<think>.*?</think>", re.I | re.DOTALL)
# Common Paddle simplified → traditional / Giản thể Paddle hay gặp → phồn thể
_S2T_PAIRS = (
    "东東丝絲两兩严嚴丧喪个個临臨为為丽麗举舉么麼义義乌烏乐樂乔喬习習乡鄉书書买買乱亂"
    "争爭于於亏虧云雲亚亞产產亩畝亲親亿億仅僅从從仑侖仓倉仪儀们們价價众眾优優会會"
    "伞傘伟偉传傳伤傷伦倫伪偽体體余餘侠俠侣侶侦偵侧側侨僑俭儉债債儿兒党黨兰蘭关關"
    "兴興养養兽獸冈岡军軍农農冯馮陈陳静靜学學给給墙牆数數门門马馬风風飞飛龙龍场場"
    "经經与與无無对對开開时時实實现現点點当當还還这這过過国國车車长長电電语語说說"
    "读讀请請谢謝见見觉覺观觀欢歡气氣汉漢万萬击擊冲衝决決况況净淨准準凉涼减減"
    "医醫听聽响響后後术術权權条條湾灣来來"
    "虽雖则則论論远遠将將难難连連"
)
_S2T_TABLE = str.maketrans(_S2T_PAIRS[0::2], _S2T_PAIRS[1::2])
# Same-meaning calligraphy variants (compare only, GT keeps written form) /
# Biến thể đồng nghĩa khi so (GT giữ đúng chữ trên ảnh)
_VARIANT_PAIRS = (
    "脩修爲為裏裡着著恒恆啓啟羣群眞真麪麵峯峰綫線牀床敎教舘館説說"
)
_VARIANT_TABLE = str.maketrans(_VARIANT_PAIRS[0::2], _VARIANT_PAIRS[1::2])
B2_FIELDS = ("image", "label", "ground_truth", "side_matter", "gemini", "post_link")
# Pure Task B2 submit (no side_matter) for /tmp local deliverables /
# Schema nộp Task B2 thuần (không side_matter) cho file local /tmp
B2_SUBMIT_FIELDS = ("image", "label", "ground_truth", "gemini", "post_link")
FLUSH_POSTS = 10
# Default pending images per run (chunk rollover) / Số ảnh pending mỗi lần chạy
DEFAULT_TARGET_CHUNK = 300
# Priority queue seq for collapsed long-line GT re-OCR /
# Seq hàng đợi ưu tiên re-OCR GT 1 dòng dài
PRIORITY_LONGLINE_SEQ = 99
# Quote re-shard seqs 80..69 so they do not collide with queues 1-12 or priority 99-88 /
# Seq shard quote 80..69, không đụng queue 1-12 hay priority 99-88
QUOTE_PARALLEL_BASE_SEQ = 80
PRIORITY_LONGLINE_QUEUE_NAME = "priority_longline.jsonl"
# Tester workbook is not Task B2 / Workbook tester không phải file nộp B2
TESTER_SCHEMA = "label-dual-tester-1.2"
# Compact HITL sheet testers fill / Sheet HITL ngắn testers điền
TESTER_HITL_FIELDS = (
    "verdict",
    "corrected_gt",
    "notes",
    "page_status",
    "glm_agree_with_fuse",
    "post_link",
    "image",
    "post_id",
    "label_caption",
    "fuse_ground_truth",
    "glm_recommend",
    "text_a",
    "text_b",
    "hitl_spans",
    "line_status_counts",
    "flags",
)
TESTER_VERDICTS = ("ok_fuse", "use_glm", "edit", "unreadable", "skip")
TESTER_REVIEW_FIELDS = (
    "page_status",
    "glm_review_bucket",
    "glm_agree_with_fuse",
    "image",
    "post_id",
    "quote_batch",
    "post_link",
    "label_caption",
    "fuse_ground_truth",
    "glm_recommend",
    "text_a",
    "text_b",
    "side_matter",
    "flags",
    "glm_flags",
    "hitl_spans",
    "line_status_counts",
    "n_lines",
    "n_hitl_spans",
    "n_locked",
    "caption_align",
    "gt_track",
    "page_conf",
    "bag",
    "cluster_bag",
    "compact_bag",
    "cer",
    "gemini_conf",
    "paddle_conf",
    "paddle_error",
    "n_ink_gemini",
    "n_ink_paddle",
    "gpt_a_conf",
    "gpt_b_conf",
    "ds_a_order",
    "ds_b_order",
    "ds_a_conf",
    "ds_b_conf",
    "glm_pick",
    "glm_source",
    "glm_conf",
    "glm_gate_ok",
    "glm_vote_score",
    "glm_dropped_invented",
    "glm_retried",
    "rec_vs_fuse_bag",
    "rec_vs_fuse_cer",
    "rec_vs_a_bag",
    "rec_vs_b_bag",
    "gemini_ink",
    "paddle_ink",
    "gemini_boxes_json",
    "paddle_boxes_json",
    "seals_json",
    "schema_version",
    "glm_schema",
    "glm_run_id",
    "page_at",
    "glm_at",
)
TESTER_LINE_FIELDS = (
    "image",
    "post_id",
    "page_status",
    "line_id",
    "status",
    "fuse",
    "text_a",
    "text_b",
    "line_conf",
    "cer",
    "bag",
    "ratio",
    "hitl_reason",
)
# Meanings for sheet review (same order as TESTER_REVIEW_FIELDS) /
# Ý nghĩa cột sheet review (cùng thứ tự TESTER_REVIEW_FIELDS)
TESTER_REVIEW_COLUMN_DOCS: tuple[tuple[str, str, str], ...] = (
    ("page_status", "Cổng tự động: silver = A≈B đủ sạch; needs_hitl = cần người xem.", "Testers: ưu tiên needs_hitl."),
    ("glm_review_bucket", "Nhóm chất lượng nháp GLM: review_ok / review_mix / review_weak / review_empty.", "Không phải trạng thái nộp B2."),
    ("glm_agree_with_fuse", "TRUE nếu GLM trùng fuse (sau S2T/biến thể). FALSE = hai bản khác nhau.", "Lọc FALSE để test trước."),
    ("image", "Đường dẫn ảnh trên MinIO (thường /images/{post_id}/{n}.jpg).", "Khóa khớp dòng; một bài có thể nhiều ảnh."),
    ("post_id", "ID bài Facebook.", "Cùng post_id = cùng bài, khác file ảnh."),
    ("quote_batch", "Queue quote 1–12 đã chia sẵn.", "Chỉ để biết batch chạy."),
    ("post_link", "Link permalink Facebook (bấm ra ảnh gốc).", "Bắt buộc mở ảnh khi chấm."),
    ("label_caption", "Caption Facebook = cột B2 `label`. Không phải mực trên ảnh.", "Đừng lấy caption làm GT thư pháp."),
    ("fuse_ground_truth", "GT fuse (Gemini mặc định, Paddle khi trùng, S2T). = B2 `ground_truth` nộp.", "Cột cần chấm: đúng mực không?"),
    ("glm_recommend", "Nháp GLM nhìn ảnh + A∪B. Không ghi vào B2, không auto-silver.", "Gợi ý thôi; chỉ dùng nếu khớp nét và chữ có trong text_a/text_b."),
    ("text_a", "OCR nhánh A (Gemini, sau GPT/DeepSeek).", "Đối chiếu glyph với fuse/GLM."),
    ("text_b", "OCR nhánh B (Paddle, sau GPT/DeepSeek). Dòng chỉ-B bị fuse bỏ.", "Khi lệch A, xem Paddle có đúng nét hơn không."),
    ("side_matter", "Chữ ngoài thân thư pháp: lạc khoản, ấn, in, UI.", "Không nhét vào ground_truth."),
    ("flags", "Cờ trang: glyph_mismatch, a_only, paddle_empty, ui_chrome, …", "Giải thích vì sao needs_hitl."),
    ("glm_flags", "Cờ log GLM: glm_gate_ok, glm_dropped_invented, glm_fallback, …", "Kỹ thuật; tester có thể bỏ."),
    ("hitl_spans", "Chỗ A≠B (glyph / a_only / partial / repeat).", "Đúng đoạn cần nhìn trên ảnh."),
    ("line_status_counts", "Đếm status từng dòng đã dóng (match, glyph_mismatch, a_only, …).", "Nhiều mismatch → đọc kỹ."),
    ("n_lines", "Số dòng thư pháp đã dóng A/B.", ""),
    ("n_hitl_spans", "Số đoạn lệch A/B.", "0 thường dễ hơn."),
    ("n_locked", "Số dòng A=B (hoặc caption vote) đã khóa.", ""),
    ("caption_align", "Caption vs A∪B: quote / partial / empty / …", "quote chỉ nghĩa caption có CJK trùng OCR, không = đúng mực."),
    ("gt_track", "Nguồn GT: fused (bình thường) hoặc none (ảnh yếu/UI).", "none → GT rỗng là đúng."),
    ("page_conf", "Confidence trung bình các dòng đã dóng (0–1).", "Thấp → nghi ngờ."),
    ("bag", "Bag Dice A vs B sau refine (0–1). Silver cần ≥ 0.75.", "Thấp = hai OCR khác túi chữ."),
    ("cluster_bag", "Bag Dice cột hình học Gemini vs Paddle (trước refine).", ""),
    ("compact_bag", "Bag Dice A vs B khi ghép hết câu (bỏ xuống dòng).", "Cao nhưng bag thấp → lệch thứ tự/cắt câu."),
    ("cer", "CER A vs B (thấp = giống). Không so với caption.", ""),
    ("gemini_conf", "Confidence trang Gemini vision.", ""),
    ("paddle_conf", "Confidence trang Paddle.", ""),
    ("paddle_error", "Lỗi HTTP/Paddle nếu có.", "Có lỗi → text_b thường trống."),
    ("n_ink_gemini", "Số box mực Gemini.", "0 = không thấy thư pháp."),
    ("n_ink_paddle", "Số box mực Paddle.", ""),
    ("gpt_a_conf", "Confidence GPT refine nhánh A.", ""),
    ("gpt_b_conf", "Confidence GPT refine nhánh B.", ""),
    ("ds_a_order", "Thứ tự đọc DeepSeek nhánh A (rtl_columns/…/unknown).", "unknown + eval_skip = giữ thứ tự GPT."),
    ("ds_b_order", "Thứ tự đọc DeepSeek nhánh B.", ""),
    ("ds_a_conf", "Confidence DeepSeek A (0 nếu skip).", ""),
    ("ds_b_conf", "Confidence DeepSeek B.", ""),
    ("glm_pick", "GLM khai pick a / b / neither.", "Không tự thành GT."),
    ("glm_source", "Nháp từ model, retry, hoặc fallback A/B.", "fallback = GLM fail, lấy nguyên một nhánh."),
    ("glm_conf", "Confidence GLM tự khai (0–1).", ""),
    ("glm_gate_ok", "TRUE nếu recommend chỉ dùng chữ trong A∪B.", "FALSE → đã cắt chữ bịa hoặc rỗng."),
    ("glm_vote_score", "Điểm nội bộ log GLM (không đổi GT).", "Kỹ thuật."),
    ("glm_dropped_invented", "TRUE nếu GLM bịa chữ rồi bị cổng cắt.", "Không tin recommend."),
    ("glm_retried", "TRUE nếu gọi GLM lần 2 vì lần 1 rỗng/lỗi.", ""),
    ("rec_vs_fuse_bag", "Bag Dice GLM vs fuse.", "1 = hai cột giống nhau."),
    ("rec_vs_fuse_cer", "CER fuse → GLM.", ""),
    ("rec_vs_a_bag", "Bag Dice GLM vs text_a.", "Cao = GLM nghiêng Gemini."),
    ("rec_vs_b_bag", "Bag Dice GLM vs text_b.", "Cao = GLM nghiêng Paddle."),
    ("gemini_ink", "Text box mực Gemini (trước GPT).", "So nếu refine làm lạ."),
    ("paddle_ink", "Text box mực Paddle (trước GPT).", ""),
    ("gemini_boxes_json", "JSON box Gemini (kèm bbox). Debug.", "Tester bỏ qua."),
    ("paddle_boxes_json", "JSON box Paddle. Debug.", "Tester bỏ qua."),
    ("seals_json", "JSON ấn/triền. Debug.", "Tester bỏ qua."),
    ("schema_version", "Schema page fuse (label-dual-1.6.0).", ""),
    ("glm_schema", "Schema log GLM.", ""),
    ("glm_run_id", "ID lần chạy GLM.", ""),
    ("page_at", "Timestamp ghi page JSON (UTC).", ""),
    ("glm_at", "Timestamp ghi GLM (UTC).", ""),
)
# Extra columns only on sheet hitl / Cột chỉ có trên sheet hitl
TESTER_HITL_COLUMN_DOCS: tuple[tuple[str, str, str], ...] = (
    ("verdict", "Tester điền: ok_fuse | use_glm | edit | unreadable | skip.", "Bắt buộc khi chấm."),
    ("corrected_gt", "GT đúng nếu verdict=edit. Mỗi câu một dòng.", "Để trống nếu ok_fuse / use_glm."),
    ("notes", "Ghi chú tự do (mờ, cắt cột, ấn lẫn mực, …).", "Không bắt buộc."),
)
VISION_MODEL = "gemini-3.6-flash-high"
GPT_MODEL = "gpt-5.6-luna"
# Empty = skip DeepSeek permute (hvb path: Gemini∥Paddle→GPT→fuse→GLM) /
# Rỗng = bỏ DeepSeek permute (luồng hvb: Gemini∥Paddle→GPT→fuse→GLM)
EVAL_MODEL = ""
PADDLE_URL = "http://paddle-ocr:8080/ocr"
# Match in-cluster Paddle max side so polygons map to pixels /
# Khớp cạnh dài Paddle trong cluster để polygon khớp pixel
PADDLE_MAX_SIDE = 2048
# Wait in Paddle FIFO; 12 jobs must not HTTP-timeout at 180s /
# Chờ hàng Paddle; 12 batch không được timeout HTTP 180s
PADDLE_TIMEOUT_SEC = 600
# Client inflight = STS replica count (2 GPU) / Inflight client = số replica STS (2 GPU)
PADDLE_MAX_INFLIGHT = 2
VOTE_SCORE_OK = 0.75
QUOTE_BATCH_COUNT = 12
# Serialize MinIO upserts from in-process batch workers /
# Khóa ghi MinIO khi worker trong cùng process
_UPSERT_LOCK = threading.Lock()
_PADDLE_SEM: threading.Semaphore | None = None
_PADDLE_SEM_LOCK = threading.Lock()
GLM_PROMPT = """You see the FULL page of Chinese calligraphy (couplet / verses).
Two OCR systems transcribed the same ink.

A (Gemini), one verse per line:
{a}

B (Paddle), one verse per line:
{b}

Write recommend_ground_truth for a human reviewer.
Rules:
1. You MUST always fill recommend_ground_truth. Never leave it empty.
2. Align columns and sentences: one complete verse per line.
3. At each character, choose the glyph from A or from B that matches the BRUSH STROKES.
4. You MUST NOT invent any character that is not in A and not in B.
5. You MUST NOT concatenate A then B (or B then A).
6. confidence is 0..1 for this recommend.

Reply JSON only, no markdown:
{{"pick":"a","recommend_ground_truth":"句一\\n句二","confidence":0.0}}
pick is a, b, or neither.
This is not final ground_truth.
"""
GLM_RETRY_PROMPT = """REQUIRED: recommend_ground_truth must be a non-empty JSON string.
One verse per line using \\n. Use only characters from A or B. Do not invent.
A:
{a}
B:
{b}
JSON only: {{"pick":"a","recommend_ground_truth":"...","confidence":0.0}}
"""

GEMINI_PROMPT = """You OCR ONE Chinese calligraphy photo.

Return ONLY compact JSON (no markdown):
{
  "confidence": 0.0,
  "gemini": [
    {
      "text": "characters in this region",
      "bounding_box": [ymin, xmin, ymax, xmax],
      "kind": "ink_text",
      "confidence": 0.0
    }
  ]
}

kind (required):
- ink_text: main handwritten calligraphy body only
- margin: handwritten colophon / 年/月/日 / small side column
- seal: stamp/chop (any color). Unreadable → text="[seal]"
- printed: printed grid, folio, cell labels
- other: UI, watermark, landscape, not writing

Rules:
- bounding_box is 0-1000 normalized [ymin, xmin, ymax, xmax]
- Traditional Chinese as written; do not simplify; do not translate
- Do NOT invent characters. Unsure glyph → text="[unreadable]" and confidence<=0.45
- No readable Chinese calligraphy → gemini=[] (or only seal/printed/other), confidence<=0.3
- ONE ink_text object per vertical COLUMN (all characters in that column, top-to-bottom concatenated). Not one object per character.
- Seals, grids, UI chrome, screenshots, watermarks → kind=seal/printed/other, never ink_text
- Array order: columns RIGHT-to-LEFT
- Traditional Chinese as written; do not simplify; do not translate
- confidence 0..1 for the page and each box
"""

GPT_REFINE_PROMPT = """You refine OCR COLUMN lines of Chinese calligraphy.
Each input string is already one vertical column (top-to-bottom).
You may ONLY: (1) reorder columns right-to-left, (2) split a column into verse lines using the SAME characters.
Do NOT emit one character per line.
Do NOT invent, replace, simplify, or add characters.
Do NOT copy Facebook captions or UI text.

Return ONLY JSON:
{{"confidence": 0.0, "layout": "rtl_columns|ltr_columns|grid|unknown", "calligraphy_lines": ["..."], "flags": []}}

Input columns (JSON):
{lines_json}
"""

DS_EVAL_PROMPT = """You judge calligraphy line order and sentence order.
You may ONLY permute the given lines. Do NOT invent characters.
Do NOT copy Facebook captions. No markdown. No <think>.
First character of the reply MUST be {{ .

Return ONLY JSON:
{{"confidence": 0.0, "reading_order": "rtl_columns|ltr_columns|grid|unknown", "calligraphy_lines": ["..."], "order_ok": true, "flags": []}}

Lines (JSON):
{lines_json}
"""


def _int_env(name: str, default: int) -> int:
    """Parse int env with fallback / Đọc env số nguyên, fallback nếu lỗi."""
    raw = os.environ.get(name, "").strip()
    if not raw:
        return default
    try:
        return int(raw)
    except ValueError:
        return default


def _bool_env(name: str, default: bool = False) -> bool:
    """Parse bool env / Đọc env boolean."""
    raw = os.environ.get(name, "").strip().lower()
    if not raw:
        return default
    return raw in {"1", "true", "yes", "on"}


def _clamp_conf(value: Any, default: float = 0.0) -> float:
    """Clamp confidence to [0, 1] / Kẹp confidence về [0, 1]."""
    try:
        num = float(value)
    except (TypeError, ValueError):
        num = default
    return round(max(0.0, min(1.0, num)), 4)


def label_dual_root(source_prefix: str, group_id: str) -> str:
    """Isolated MinIO prefix for this pilot / Prefix MinIO tách cho pilot."""
    return f"{_group_root(source_prefix, group_id)}/ocr/label_dual_pilot"


def facebook_post_link(
    group_id: str,
    post_id: str,
    existing: str = "",
) -> str:
    """Clickable Facebook group permalink / Permalink group Facebook để click.

    Prefer a crawl URL that already points at this post; else canonical
    ``/groups/{gid}/permalink/{pid}/``.
    Ưu tiên URL crawl đã trỏ đúng bài; không có thì URL permalink chuẩn.
    """
    text = str(existing or "").strip()
    pid = str(post_id or "").strip()
    low = text.lower()
    if text.startswith("http") and "facebook.com" in low:
        if not pid or pid in text:
            return text
    gid = str(group_id or "").strip()
    if gid and pid and pid not in {"unknown", ""}:
        return permalink_for(gid, pid)
    return text


def _assert_pilot_write_root(root: str) -> str:
    """Refuse writes outside ocr/label_dual_pilot / Cấm ghi ngoài ocr/label_dual_pilot."""
    text = str(root or "").strip().strip("/")
    if "/ocr/label_dual_pilot" not in f"/{text}":
        raise RuntimeError(f"refuse write outside label_dual_pilot: {root}")
    return text


def levenshtein(left: str, right: str) -> int:
    """Edit distance / Khoảng cách sửa."""
    a, b = str(left or ""), str(right or "")
    if a == b:
        return 0
    if not a:
        return len(b)
    if not b:
        return len(a)
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        cur = [i]
        for j, cb in enumerate(b, 1):
            cur.append(
                min(cur[-1] + 1, prev[j] + 1, prev[j - 1] + (0 if ca == cb else 1))
            )
        prev = cur
    return prev[-1]


def cer(ref: str, hyp: str) -> float:
    """Character error rate on CJK compact text / CER trên chuỗi CJK compact."""
    a, b = norm_cjk(ref), norm_cjk(hyp)
    if not a and not b:
        return 0.0
    if not a:
        return 1.0
    return round(levenshtein(a, b) / len(a), 4)


def wer_lines(ref_lines: list[str], hyp_lines: list[str]) -> float:
    """WER treating each line as a token / WER mỗi dòng là một token."""
    ref = [norm_cjk(x) for x in ref_lines if norm_cjk(x)]
    hyp = [norm_cjk(x) for x in hyp_lines if norm_cjk(x)]
    if not ref and not hyp:
        return 0.0
    if not ref:
        return 1.0
    return round(levenshtein_tokens(ref, hyp) / len(ref), 4)


def levenshtein_tokens(ref: list[str], hyp: list[str]) -> int:
    """Edit distance on token lists / Edit distance trên list token."""
    if ref == hyp:
        return 0
    if not ref:
        return len(hyp)
    if not hyp:
        return len(ref)
    prev = list(range(len(hyp) + 1))
    for i, ra in enumerate(ref, 1):
        cur = [i]
        for j, hb in enumerate(hyp, 1):
            cur.append(min(cur[-1] + 1, prev[j] + 1, prev[j - 1] + (0 if ra == hb else 1)))
        prev = cur
    return prev[-1]


def fuzz_ratio(left: str, right: str) -> float:
    """0–100 similarity / Độ giống 0–100."""
    a, b = norm_cjk(left), norm_cjk(right)
    if not a and not b:
        return 100.0
    if not a or not b:
        return 0.0
    return round(SequenceMatcher(None, a, b, autojunk=False).ratio() * 100.0, 2)


def bag_dice(left: str, right: str) -> float:
    """Char-bag Dice (order-insensitive) / Dice túi chữ (không cần thứ tự)."""
    a, b = norm_cjk(left), norm_cjk(right)
    if not a and not b:
        return 1.0
    ca, cb = Counter(a), Counter(b)
    inter = sum((ca & cb).values())
    tot = sum(ca.values()) + sum(cb.values())
    return round(2.0 * inter / tot, 4) if tot else 1.0


def partial_ratio(left: str, right: str) -> float:
    """Substring-friendly similarity 0–100 / Giống substring 0–100."""
    a, b = norm_cjk(left), norm_cjk(right)
    if not a and not b:
        return 100.0
    if not a or not b:
        return 0.0
    short, long = (a, b) if len(a) <= len(b) else (b, a)
    block = SequenceMatcher(None, short, long, autojunk=False).find_longest_match(
        0, len(short), 0, len(long)
    )
    return round(100.0 * int(block.size) / max(len(short), 1), 2)


def invented_chars(raw: str, refined: str) -> bool:
    """True if refined added CJK not in raw / True nếu refined thêm CJK không có trong raw."""
    src, dst = compact_cjk(raw), compact_cjk(refined)
    if not dst:
        return False
    return bool(set(dst) - set(src))


def bbox_of(box: dict[str, Any]) -> list[int]:
    """Normalize [ymin,xmin,ymax,xmax] to 0–1000 / Chuẩn hoá bbox về 0–1000."""
    raw = box.get("bounding_box") or box.get("bbox") or [0, 0, 0, 0]
    if not isinstance(raw, list) or len(raw) < 4:
        return [0, 0, 0, 0]
    try:
        return [max(0, min(1000, int(round(float(x))))) for x in raw[:4]]
    except (TypeError, ValueError):
        return [0, 0, 0, 0]


def poly_to_1000(poly: Any, width: int, height: int) -> list[int]:
    """Paddle polygon → 0–1000 bbox / Polygon Paddle → bbox 0–1000."""
    points: list[tuple[float, float]] = []
    if isinstance(poly, (list, tuple)):
        for pt in poly:
            if isinstance(pt, (list, tuple)) and len(pt) >= 2:
                points.append((float(pt[0]), float(pt[1])))
    if not points or width <= 0 or height <= 0:
        return [0, 0, 0, 0]
    xs = [p[0] for p in points]
    ys = [p[1] for p in points]
    ymin = 1000.0 * min(ys) / height
    xmin = 1000.0 * min(xs) / width
    ymax = 1000.0 * max(ys) / height
    xmax = 1000.0 * max(xs) / width
    return [max(0, min(1000, int(round(v)))) for v in (ymin, xmin, ymax, xmax)]


def classify_paddle_block(block: dict[str, Any], *, img_w: int, img_h: int) -> str:
    """Heuristic kind for Paddle boxes / Kind heuristic cho box Paddle."""
    text = str(block.get("text") or "")
    compact = "".join(ch for ch in text if not ch.isspace())
    cjk = len(CJK_RE.findall(compact))
    latin = len(LATIN_RE.findall(compact))
    total = max(len(compact), 1)
    cjk_ratio = cjk / total
    bb = bbox_of(block) if block.get("bounding_box") else poly_to_1000(block.get("box"), img_w, img_h)
    ymin, xmin, ymax, xmax = bb
    bw, bh = max(xmax - xmin, 1), max(ymax - ymin, 1)
    area = bw * bh
    if bh > bw * 1.15 and cjk_ratio >= 0.25:
        return "ink_text"
    if abs(bw - bh) / max(bw, bh) < 0.35 and area < 40_000 and cjk_ratio < 0.35:
        return "seal"
    if cjk_ratio < 0.2 and (latin / total >= 0.4 or not compact):
        return "printed" if compact else "other"
    if cjk_ratio >= 0.25:
        return "ink_text"
    return "other"


def rtl_key(box: dict[str, Any]) -> tuple[int, int]:
    """Sort key: columns right-to-left, then top-to-bottom.

    Khóa sort: cột phải→trái, rồi trên→dưới.
    """
    bb = bbox_of(box)
    return (-bb[3], bb[0])


def grid_key(box: dict[str, Any]) -> tuple[int, int]:
    bb = bbox_of(box)
    return (bb[0], bb[1])


def boxes_to_lines(boxes: list[dict[str, Any]], layout: str) -> list[str]:
    """Join calligraphy boxes into page lines / Ghép box thư pháp thành dòng trang."""
    key = grid_key if layout == "grid" else rtl_key
    ordered = sorted(boxes, key=key)
    return [str(b.get("text") or "").strip() for b in ordered if str(b.get("text") or "").strip()]


def to_traditional(text: str) -> str:
    """Map common simplified glyphs to traditional / Đổi giản thể thường gặp sang phồn."""
    return str(text or "").translate(_S2T_TABLE)


def fold_cjk_variants(text: str) -> str:
    """Fold equivalent calligraphy glyphs for compare only.

    Gộp biến thể thư pháp đồng nghĩa (chỉ khi so, không ghi GT).
    """
    return str(text or "").translate(_VARIANT_TABLE)


def norm_cjk(text: str) -> str:
    """CJK compact + S2T + variant fold for metrics / CJK gọn + S2T + biến thể để chấm."""
    return fold_cjk_variants(compact_cjk(to_traditional(text)))


def latin_heavy(text: str) -> bool:
    """True when Latin letters dominate — screenshot/print, not ink.

    True khi chữ Latin chiếm nhiều — ảnh chụp/in, không phải mực.
    """
    compact = "".join(ch for ch in str(text or "") if not ch.isspace())
    latin = len(LATIN_RE.findall(compact))
    cjk = len(CJK_RE.findall(compact))
    # Six+ Latin letters at ≥40% of CJK, or 8+ letters regardless /
    # ≥6 chữ Latin và ≥40% so với CJK, hoặc ≥8 chữ Latin
    if latin >= 6 and latin >= 0.4 * max(cjk, 1):
        return True
    return latin >= 8


def strip_caption_noise(text: str) -> str:
    """Drop URL, hashtag, mention, non-CJK from a Facebook caption.

    Bỏ URL, hashtag, mention, ký tự không CJK khỏi caption Facebook.
    """
    raw = _CAPTION_URL_RE.sub(" ", str(text or ""))
    raw = _CAPTION_HASH_RE.sub(" ", raw)
    raw = _CAPTION_MENTION_RE.sub(" ", raw)
    kept = [ch if CJK_RE.match(ch) or ch.isspace() else " " for ch in raw]
    return re.sub(r"\s+", " ", "".join(kept)).strip()


def caption_use(caption: str, allowed: set[str]) -> tuple[str, str]:
    """Classify caption for fuse; return (align, cjk_hint).

    Phân loại caption cho fuse; trả (align, CJK gợi ý).
    """
    cjk = compact_cjk(to_traditional(strip_caption_noise(caption)))
    if not cjk:
        return "empty", ""
    if not allowed:
        return "unrelated", ""
    inter = sum(1 for ch in cjk if ch in allowed)
    if inter >= 4 or inter / max(len(cjk), 1) >= 0.5:
        return "quote", cjk
    return "unrelated", ""


def caption_pick(text_a: str, text_b: str, caption_cjk: str) -> str | None:
    """Vote one 1:1 glyph using caption CJK already in A∪B.

    Vote 1 glyph 1:1 bằng CJK caption đã nằm trong A∪B.
    """
    if not caption_cjk:
        return None
    a = compact_cjk(to_traditional(text_a))
    b = compact_cjk(to_traditional(text_b))
    if a == b or len(a) != len(b):
        return None
    diffs = [(ca, cb) for ca, cb in zip(a, b) if ca != cb]
    if len(diffs) != 1:
        return None
    ca, cb = diffs[0]
    in_a = ca in caption_cjk
    in_b = cb in caption_cjk
    if in_a and not in_b:
        return to_traditional(text_a)
    if in_b and not in_a:
        return to_traditional(text_b)
    return None


def dominant_repeat_mismatch(text_a: str, text_b: str) -> bool:
    """True when a frequent glyph has different counts on A vs B.

    True khi chữ lặp nhiều nhưng hai nhánh đếm khác nhau.
    """
    ca = Counter(compact_cjk(to_traditional(text_a)))
    cb = Counter(compact_cjk(to_traditional(text_b)))
    n = max(sum(ca.values()), sum(cb.values()), 1)
    for ch, cnt in (ca | cb).most_common():
        if cnt / n >= 0.4 and ca[ch] != cb[ch]:
            return True
    return False


def invented_multiset(raw: str, refined: str) -> bool:
    """True if refined added CJK counts not in raw / True nếu refined tăng count CJK."""
    src = Counter(compact_cjk(raw))
    dst = Counter(compact_cjk(refined))
    for ch, n in dst.items():
        if n > src[ch]:
            return True
    return False


def accept_ab_recommend(text_a: str, text_b: str, recommend: str) -> str:
    """Keep recommend only if it reuses A∪B glyphs; keep verse newlines.

    Chỉ giữ recommend khi tái sử dụng chữ A∪B; giữ xuống dòng câu.
    Human still reviews; empty means drop the suggestion /
    Người vẫn duyệt; rỗng = bỏ gợi ý.
    """
    a = to_traditional(str(text_a or "").strip())
    b = to_traditional(str(text_b or "").strip())
    rec_lines = [
        to_traditional(x.strip())
        for x in str(recommend or "").replace("\\n", "\n").splitlines()
        if compact_cjk(x)
    ]
    rec_join = "".join(rec_lines)
    rec_c = compact_cjk(rec_join)
    a_c, b_c = compact_cjk(a), compact_cjk(b)
    if not rec_c or not (a_c or b_c):
        return ""
    if invented_multiset(a + b, rec_join):
        return ""
    # Concat A+B only when both tracks have CJK / Chỉ cấm dán A+B khi cả hai nhánh có CJK
    if a_c and b_c and rec_c in {a_c + b_c, b_c + a_c}:
        return ""
    # Max length = multiset union of A and B / Độ dài tối đa = hợp túi chữ A và B
    union_n = sum((Counter(a_c) | Counter(b_c)).values())
    if len(rec_c) > union_n:
        return ""
    a_verses = [compact_cjk(to_traditional(x)) for x in a.splitlines() if compact_cjk(x)]
    # One blob → cut by A's verse lengths / Một khối → cắt theo độ dài câu A
    if len(rec_lines) == 1 and len(a_verses) >= 2:
        cut = _cut_join_by_template(rec_c, a_verses)
        if len(cut) == len(a_verses) and not invented_multiset(a + b, "".join(cut)):
            rec_lines = cut
    return "\n".join(rec_lines)


def compute_glm_vote(
    *,
    text_a: str,
    text_b: str,
    recommend: str,
    pick: str,
    source: str,
    glm_conf: float,
    gate_ok: bool,
    dropped_invented: bool,
) -> dict[str, Any]:
    """Quantify A/B glyph vote after the A∪B gate / Lượng hóa vote A/B sau cổng A∪B."""
    a = compact_cjk(to_traditional(text_a))
    b = compact_cjk(to_traditional(text_b))
    rec_gated = accept_ab_recommend(text_a, text_b, recommend)
    rec = compact_cjk(to_traditional(rec_gated))
    rec_in = compact_cjk(to_traditional(str(recommend or "")))
    n_pair = min(len(a), len(b))
    n_disagree = sum(1 for i in range(n_pair) if a[i] != b[i])
    n_align = min(n_pair, len(rec))
    n_pick_a = n_pick_b = n_pick_other = 0
    for i in range(n_align):
        if a[i] == b[i]:
            continue
        if rec[i] == a[i]:
            n_pick_a += 1
        elif rec[i] == b[i]:
            n_pick_b += 1
        else:
            n_pick_other += 1
    union_n = sum((Counter(a) | Counter(b)).values())
    union_cover = round(len(rec) / union_n, 4) if union_n else 0.0
    mix = bool(n_pick_a > 0 and n_pick_b > 0 and rec != a and rec != b)
    concat_reject = bool(a and b and rec_in in {a + b, b + a})
    invent_rate = 1.0 if dropped_invented else 0.0
    n_verses_a = len([x for x in str(text_a or "").splitlines() if compact_cjk(x)])
    n_verses_rec = len([x for x in str(rec_gated or "").splitlines() if compact_cjk(x)])
    verse_delta = abs(n_verses_rec - n_verses_a)
    pick_l = str(pick or "").strip().lower()
    pick_mismatch = bool(
        (pick_l == "a" and n_pick_b > n_pick_a)
        or (pick_l == "b" and n_pick_a > n_pick_b)
    )
    src = str(source or "")
    if src == "model":
        src_w = 1.0
    elif src == "retry":
        src_w = 0.5
    elif src.startswith("fallback"):
        src_w = 0.2
    else:
        src_w = 0.0
    conf = _clamp_conf(glm_conf, 0.0)
    vote_score = round(
        0.35 * (1.0 if gate_ok else 0.0)
        + 0.20 * src_w
        + 0.20 * conf
        + 0.15 * (1.0 - invent_rate)
        + 0.10 * (0.0 if pick_mismatch else 1.0),
        4,
    )
    if src == "glm_empty_ab" or (not rec and not (a or b)):
        bucket = "review_empty"
    elif (
        src.startswith("fallback")
        or conf < 0.5
        or pick_mismatch
        or verse_delta >= 2
        or not gate_ok
    ):
        bucket = "review_weak"
    elif mix and n_disagree >= 1:
        bucket = "review_mix"
    elif vote_score >= VOTE_SCORE_OK and gate_ok:
        bucket = "review_ok"
    else:
        bucket = "review_weak"
    return {
        "n_disagree": n_disagree,
        "n_pick_a": n_pick_a,
        "n_pick_b": n_pick_b,
        "n_pick_other": n_pick_other,
        "mix_rate": 1.0 if mix else 0.0,
        "union_cover": union_cover,
        "invent_rate": invent_rate,
        "concat_reject": 1.0 if concat_reject else 0.0,
        "verse_delta": verse_delta,
        "pick_mismatch": pick_mismatch,
        "vote_score": vote_score,
        "review_bucket": bucket,
    }


def _strip_llm_json(raw: str) -> str:
    """Strip think tags before JSON extract / Bỏ thẻ think trước khi parse JSON."""
    return _THINK_RE.sub("", str(raw or "")).strip()


def fuse_aligned_lines(
    aligned: list[dict[str, Any]],
    *,
    caption: str = "",
    g_cluster: list[str] | None = None,
) -> dict[str, Any]:
    """Fuse paired A/B lines into draft GT with HITL spans.

    Ghép dòng A/B đã dóng thành GT nháp kèm span HITL.
    """
    allowed: set[str] = set()
    for row in aligned:
        allowed |= set(compact_cjk(to_traditional(str(row.get("text_a") or ""))))
        allowed |= set(compact_cjk(to_traditional(str(row.get("text_b") or ""))))
    cap_align, cap_cjk = caption_use(caption, allowed)
    gt_lines: list[str] = []
    hitl_spans: list[dict[str, str]] = []
    fused: list[dict[str, Any]] = []
    locked: list[str] = []
    for row in aligned:
        ta = str(row.get("text_a") or "")
        tb = str(row.get("text_b") or "")
        st = str(row.get("status") or "")
        new = dict(row)
        if st == "b_only":
            new["fuse"] = "drop_b"
            fused.append(new)
            continue
        if st == "a_only":
            hitl_spans.append({"a": ta, "b": "", "reason": "a_only"})
            new["fuse"] = "draft_a"
            if compact_cjk(ta):
                gt_lines.append(to_traditional(ta))
            fused.append(new)
            continue
        ta_t = to_traditional(ta)
        # Compare after variant fold; keep written form in GT /
        # So sau khi gộp biến thể; GT giữ đúng chữ viết
        a_c = norm_cjk(ta)
        b_c = norm_cjk(tb)
        if a_c and a_c == b_c:
            gt_lines.append(ta_t or ta)
            locked.append(ta_t or ta)
            new["status"] = "match"
            new["fuse"] = "equal_s2t"
            fused.append(new)
            continue
        if dominant_repeat_mismatch(ta, tb):
            hitl_spans.append({"a": ta, "b": tb, "reason": "repeat_count"})
            if compact_cjk(ta_t):
                gt_lines.append(ta_t)
            new["fuse"] = "hitl_repeat"
            fused.append(new)
            continue
        picked = caption_pick(ta, tb, cap_cjk)
        if picked and not invented_multiset(ta + tb, picked):
            gt_lines.append(picked)
            locked.append(picked)
            new["status"] = "match"
            new["fuse"] = "caption_vote"
            fused.append(new)
            continue
        if b_c and a_c and b_c in a_c and len(b_c) / max(len(a_c), 1) < 0.8:
            hitl_spans.append({"a": ta, "b": tb, "reason": "partial"})
            gt_lines.append(ta_t)
            new["fuse"] = "hitl_partial"
            fused.append(new)
            continue
        hitl_spans.append({"a": ta, "b": tb, "reason": st or "glyph"})
        if compact_cjk(ta_t):
            gt_lines.append(ta_t)
        new["fuse"] = "hitl_glyph"
        fused.append(new)
    if g_cluster:
        gt_lines = prefer_split_keep_order(g_cluster, gt_lines)
    return {
        "gt_lines": gt_lines,
        "aligned": fused,
        "hitl_spans": hitl_spans,
        "locked_lines": locked,
        "caption_align": cap_align,
        "gt_track": "fused",
    }


def recompute_from_page(page: dict[str, Any], *, caption: str = "") -> dict[str, Any]:
    """Replay fuse+gate on a saved page JSON (no Gemini/Paddle).

    Chạy lại fuse+cổng trên page JSON đã OCR (không gọi Gemini/Paddle).
    """
    rows = list(page.get("calligraphy_lines") or [])
    g_eval = [str(r.get("text_a") or "") for r in rows if str(r.get("text_a") or "").strip()]
    p_eval = [str(r.get("text_b") or "") for r in rows if str(r.get("text_b") or "").strip()]
    g_before, p_before = list(g_eval), list(p_eval)
    g_eval, p_eval = recover_wrapped_lines(g_eval, p_eval)
    wrap_recovered = g_eval != g_before or p_eval != p_before
    aligned = _align_line_lists(g_eval, p_eval)
    fused = fuse_aligned_lines(aligned, caption=caption, g_cluster=g_eval)
    page_bag = bag_dice("\n".join(g_eval), "\n".join(p_eval))
    page_cer = cer("\n".join(g_eval), "\n".join(p_eval))
    align_meta = page.get("align") if isinstance(page.get("align"), dict) else {}
    cluster_bag = float(align_meta.get("cluster_bag") or page_bag)
    cluster_cer = float(align_meta.get("cluster_cer") or page_cer)
    compact_bag = bag_dice("".join(g_eval), "".join(p_eval))
    scored = [r for r in fused["aligned"] if r.get("status") not in {"a_only", "b_only"}]
    base = scored or fused["aligned"]
    page_conf = (
        _clamp_conf(sum(float(r.get("line_conf") or 0) for r in base) / len(base)) if base else 0.0
    )
    stored = page.get("boxes") if isinstance(page.get("boxes"), dict) else {}
    g_boxes = list(stored.get("gemini") or [])
    p_boxes = list(stored.get("paddle") or [])
    flags = collect_page_flags(
        [str(x) for x in (page.get("flags") or [])],
        aligned=fused["aligned"],
        fused=fused,
        wrap_recovered=wrap_recovered,
        g_eval=g_eval,
        p_eval=p_eval,
        g_boxes=g_boxes or None,
        p_boxes=p_boxes or None,
    )
    paddle = page.get("paddle") if isinstance(page.get("paddle"), dict) else {}
    p_err = paddle.get("error")
    status = decide_page_status(
        g_eval=g_eval,
        p_eval=p_eval,
        aligned=fused["aligned"],
        page_conf=page_conf,
        page_bag=page_bag,
        page_cer=page_cer,
        cluster_bag=cluster_bag,
        cluster_cer=cluster_cer,
        flags=flags,
        p_err=str(p_err) if p_err else None,
        compact_bag=compact_bag,
    )
    if "weak_evidence" in flags or "ui_chrome" in flags:
        gt_lines, track = [], "none"
    else:
        gt_lines, track = fused["gt_lines"], fused["gt_track"]
    return {
        "post_id": page.get("post_id"),
        "old_status": page.get("page_status"),
        "page_status": status,
        "ground_truth": "\n".join(to_traditional(x) for x in gt_lines if str(x).strip()),
        "gt_track": track,
        "caption_align": fused["caption_align"],
        "hitl_spans": fused["hitl_spans"],
        "locked_lines": fused["locked_lines"],
        "compact_bag": compact_bag,
        "cluster_bag": cluster_bag,
        "page_bag": page_bag,
        "page_cer": page_cer,
        "aligned": fused["aligned"],
        "printed_script": "printed_script" in flags or "latin_heavy" in flags,
        "flags": flags,
    }


def _box_cx(box: dict[str, Any]) -> float:
    bb = bbox_of(box)
    return (bb[1] + bb[3]) / 2.0


def _box_cy(box: dict[str, Any]) -> float:
    bb = bbox_of(box)
    return (bb[0] + bb[2]) / 2.0


def _box_width(box: dict[str, Any]) -> int:
    bb = bbox_of(box)
    return max(bb[3] - bb[1], 1)


def cluster_boxes_to_columns(boxes: list[dict[str, Any]]) -> list[str]:
    """Group boxes into RTL columns; join top-to-bottom inside each column.

    Gom box thành cột phải→trái; trong cột ghép trên→dưới.
    """
    items = [b for b in boxes if str(b.get("text") or "").strip()]
    if not items:
        return []
    widths = sorted(_box_width(b) for b in items)
    median_w = widths[len(widths) // 2]
    thresh = max(40.0, min(140.0, 0.7 * float(median_w)))
    # Rightmost boxes first so new clusters are new columns to the left /
    # Box phải nhất trước để cluster mới là cột bên trái
    ordered = sorted(items, key=lambda b: (-_box_cx(b), _box_cy(b)))
    columns: list[list[dict[str, Any]]] = []
    for box in ordered:
        cx = _box_cx(box)
        placed = False
        for col in columns:
            col_cx = sum(_box_cx(b) for b in col) / len(col)
            if abs(cx - col_cx) <= thresh:
                col.append(box)
                placed = True
                break
        if not placed:
            columns.append([box])
    columns.sort(key=lambda col: -sum(_box_cx(b) for b in col) / len(col))
    lines: list[str] = []
    for col in columns:
        col.sort(key=lambda b: (_box_cy(b), bbox_of(b)[0]))
        text = "".join(str(b.get("text") or "").strip() for b in col)
        if text:
            lines.append(text)
    return lines


def is_per_char_boxes(boxes: list[dict[str, Any]]) -> bool:
    """True when most ink boxes are single characters / True khi đa số box chỉ 1 chữ."""
    lens = [len(compact_cjk(str(b.get("text") or ""))) for b in boxes]
    lens = [n for n in lens if n > 0]
    if len(lens) < 4:
        return False
    lens.sort()
    return lens[len(lens) // 2] <= 1


def pick_gt_track(
    *,
    g_lines: list[str],
    p_lines: list[str],
) -> tuple[list[str], str]:
    """Choose B2 ground_truth source / Chọn nguồn ground_truth cho B2."""
    if not p_lines:
        return g_lines, "gemini"
    if not g_lines:
        return p_lines, "paddle"
    bag = bag_dice("\n".join(g_lines), "\n".join(p_lines))
    g_lens = [len(compact_cjk(x)) for x in g_lines if compact_cjk(x)]
    p_lens = [len(compact_cjk(x)) for x in p_lines if compact_cjk(x)]
    g_med = sorted(g_lens)[len(g_lens) // 2] if g_lens else 0
    p_med = sorted(p_lens)[len(p_lens) // 2] if p_lens else 0
    # Same chars but Gemini still per-character lines / Cùng chữ nhưng Gemini còn từng chữ một dòng
    if bag >= 0.9 and g_med <= 1 and p_med >= 2:
        return p_lines, "paddle"
    if bag >= 0.9 and p_med >= 4 and g_med <= 2 and len(p_lines) >= 2:
        return p_lines, "paddle"
    return g_lines, "gemini"


def keep_overlapping_lines(
    anchor: list[str],
    other: list[str],
    *,
    min_frac: float = 0.5,
) -> list[str]:
    """Drop other-lines with little CJK overlap vs the anchor page.

    Bỏ dòng other ít trùng CJK với trang neo (lạc khoản/ấn Paddle đọc nhầm).
    """
    src = set(compact_cjk("".join(anchor)))
    if not src:
        return list(other)
    kept: list[str] = []
    for line in other:
        chars = compact_cjk(line)
        if not chars:
            continue
        frac = sum(1 for ch in chars if ch in src) / len(chars)
        if frac >= min_frac:
            kept.append(line)
    return kept


def prefer_split_keep_order(cluster: list[str], refined: list[str]) -> list[str]:
    """Keep verse split if concat order matches; else keep geometric cluster order.

    Giữ cắt câu nếu cùng thứ tự ghép; không thì giữ thứ tự cột hình học.
    Cluster lines are S2T'd so simplified OCR cannot overwrite traditional GT /
    Cụm dòng được S2T để OCR giản thể không ghi đè GT phồn thể.
    """
    if not refined:
        return [to_traditional(x) for x in cluster]
    if not cluster:
        return [to_traditional(x) for x in refined]
    a_raw = compact_cjk("".join(cluster))
    b_raw = compact_cjk("".join(refined))
    a = norm_cjk("".join(cluster))
    b = norm_cjk("".join(refined))
    # Identical concat (already same script) → keep verse split /
    # Ghép giống hệt (cùng kiểu chữ) → giữ cắt câu
    if a_raw == b_raw:
        return list(refined)
    # Same bag after S2T but different script/order → cluster order + traditional /
    # Cùng túi chữ sau S2T nhưng khác kiểu/thứ tự → thứ tự cụm + phồn thể
    if a == b or bag_dice(a, b) >= 0.98:
        return [to_traditional(x) for x in cluster]
    return [to_traditional(x) for x in refined]


def _cut_join_by_template(joined: str, templates: list[str]) -> list[str]:
    """Slice joined CJK by template line lengths / Cắt CJK đã ghép theo độ dài mẫu."""
    templates = [t for t in templates if t]
    if not joined or not templates:
        return []
    out: list[str] = []
    i = 0
    for k, tmpl in enumerate(templates):
        n = len(tmpl)
        if k == len(templates) - 1:
            piece = joined[i:]
        else:
            piece = joined[i : i + n]
            i += n
        if not piece:
            return []
        out.append(piece)
    return out


def recover_wrapped_lines(g_lines: list[str], p_lines: list[str]) -> tuple[list[str], list[str]]:
    """Re-cut wrapped columns when concat sequence matches; no new glyphs.

    Cắt lại cột bị dính khi chuỗi ghép khớp; không thêm glyph.
    Requires high sequence fuzz so scrambled fragments are not reordered /
    Cần fuzz chuỗi cao để không sắp lại mảnh Paddle lộn xộn.
    """
    g = [x for x in g_lines if compact_cjk(x)]
    p = [x for x in p_lines if compact_cjk(x)]
    if not g or not p or len(g) == len(p):
        return g_lines, p_lines
    g_join, p_join = norm_cjk("".join(g)), norm_cjk("".join(p))
    if bag_dice(g_join, p_join) < 0.92 or fuzz_ratio(g_join, p_join) < 88.0:
        return g_lines, p_lines
    if len(p) < len(g) and len(g) >= 2:
        cut = _cut_join_by_template(p_join, [norm_cjk(x) for x in g])
        if len(cut) == len(g):
            return g, cut
    if len(g) < len(p) and len(p) >= 2:
        cut = _cut_join_by_template(g_join, [norm_cjk(x) for x in p])
        if len(cut) == len(p):
            return cut, p
    return g_lines, p_lines


def line_conf_metric(cer_v: float, ratio: float) -> float:
    return _clamp_conf(0.5 * (1.0 - cer_v) + 0.5 * (ratio / 100.0))


def gate_line(*, cer_v: float, ratio: float, bag: float, partial: float, rec: float) -> str:
    """Line agreement status / Trạng thái khớp một dòng."""
    if cer_v <= 0.08 and ratio >= 92 and rec >= 0.85:
        return "match"
    if bag >= 0.92 and cer_v > 0.08:
        return "order_fix"
    if partial >= 90 and rec < 0.7:
        return "partial"
    if cer_v > 0.20 and bag < 0.80:
        return "glyph_mismatch"
    return "weak"


def paired_line_statuses(aligned: list[dict[str, Any]]) -> list[str]:
    """Statuses of Gemini–Paddle paired rows / Status các dòng đã dóng hai nhánh."""
    return [
        str(row.get("status") or "")
        for row in aligned
        if str(row.get("status") or "") not in {"a_only", "b_only"}
    ]


def _ink_confs(boxes: list[dict[str, Any]]) -> list[float]:
    """Ink-box confidences >0 / Conf các box mực >0."""
    out: list[float] = []
    for box in boxes:
        if str(box.get("kind") or "") != "ink_text":
            continue
        conf = _clamp_conf(box.get("confidence"), 0.0)
        if conf > 0:
            out.append(conf)
    return out


def collect_page_flags(
    flags: list[str],
    *,
    aligned: list[dict[str, Any]],
    fused: dict[str, Any] | None = None,
    wrap_recovered: bool = False,
    g_eval: list[str] | None = None,
    p_eval: list[str] | None = None,
    g_boxes: list[dict[str, Any]] | None = None,
    p_boxes: list[dict[str, Any]] | None = None,
) -> list[str]:
    """Promote line/OCR signals onto page flags; drop status duplicates.

    Đưa tín hiệu dòng/OCR lên flag trang; bỏ flag trùng status.
    """
    out = [str(x) for x in flags if str(x) not in DROP_PAGE_FLAGS]
    statuses = [str(row.get("status") or "") for row in aligned]
    paired = [s for s in statuses if s not in {"a_only", "b_only", ""}]
    if "glyph_mismatch" in statuses or "weak" in statuses:
        out.append("glyph_mismatch")
    if "partial" in statuses:
        out.append("partial")
    if "a_only" in statuses:
        out.append("a_only")
    if "b_only" in statuses:
        out.append("b_only_dropped")
    fuse_names = {str(row.get("fuse") or "") for row in aligned}
    span_reasons = {str(s.get("reason") or "") for s in (fused or {}).get("hitl_spans") or []}
    if "hitl_repeat" in fuse_names or "repeat_count" in span_reasons:
        out.append("repeat_count")
    if paired and all(s in {"match", "order_fix"} for s in paired) and "order_fix" in paired:
        if not ({"glyph_mismatch", "partial", "a_only", "repeat_count"} & set(out)):
            out.append("order_fix_only")
    if wrap_recovered:
        out.append("wrap_recovered")
    if "caption_vote" in fuse_names:
        out.append("caption_vote")
    g_txt = "".join(g_eval or [])
    p_txt = "".join(p_eval or [])
    if latin_heavy(g_txt) or latin_heavy(p_txt):
        out.append("latin_heavy")
        out.append("printed_script")
    g_given = g_boxes is not None
    p_given = p_boxes is not None
    g_boxes = list(g_boxes or [])
    p_boxes = list(p_boxes or [])
    if any(str(b.get("kind") or "") == "printed" and compact_cjk(str(b.get("text") or "")) for b in g_boxes + p_boxes):
        out.append("printed_script")
    if g_given and not any(str(b.get("kind") or "") == "ink_text" for b in g_boxes):
        out.append("gemini_empty_ink")
    ink_confs = _ink_confs(g_boxes) + _ink_confs(p_boxes)
    if ink_confs and min(ink_confs) < 0.45:
        out.append("low_box_conf")
    if "ui_chrome" in out:
        out.append("screenshot")
    elif "latin_heavy" in out and any(str(b.get("kind") or "") == "other" for b in g_boxes + p_boxes):
        out.append("screenshot")
    return sorted(set(out))


def decide_page_status(
    *,
    g_eval: list[str],
    p_eval: list[str],
    aligned: list[dict[str, Any]],
    page_conf: float,
    page_bag: float,
    page_cer: float,
    cluster_bag: float,
    cluster_cer: float,
    flags: list[str],
    p_err: str | None,
    compact_bag: float | None = None,
) -> str:
    """Silver only if paired lines are clean and compact cover is enough.

    Silver chỉ khi dòng đã dóng sạch và phủ compact đủ (không phạt 1 vs 2 cột).
    """
    if set(flags) & BLOCK_SILVER_FLAGS:
        return "needs_hitl"
    if "weak_evidence" in flags or "ui_chrome" in flags or "printed_script" in flags:
        return "needs_hitl"
    # Raw Latin (ignored by CJK bag) means screenshot/print, not ink /
    # Chữ Latin thô (bag CJK bỏ qua) = ảnh chụp/in, không phải mực
    if latin_heavy("".join(g_eval)) or latin_heavy("".join(p_eval)):
        return "needs_hitl"
    if not g_eval:
        return "needs_hitl"
    paired = paired_line_statuses(aligned)
    # Every paired line must match; weak/glyph cannot ride the page bag shortcut /
    # Mọi dòng đã dóng phải khớp; weak/glyph không được đi cửa bag cả trang
    lines_clean = bool(paired) and all(status in {"match", "order_fix"} for status in paired)
    n_a = max(len([row for row in aligned if row.get("text_a")]), 1)
    match_n = sum(1 for row in aligned if row.get("status") in {"match", "order_fix"})
    cover_src = page_bag if compact_bag is None else compact_bag
    # Compact bag covers Gemini 1 col vs Paddle 2 with the same chars /
    # Bag compact phủ Gemini 1 cột vs Paddle 2 cùng chữ
    cover_ok = cluster_bag >= 0.75 or cover_src >= 0.75
    silver = False
    if lines_clean and match_n / n_a >= 0.8 and page_conf >= 0.65 and cover_ok:
        silver = True
    if lines_clean and p_eval and page_bag >= 0.92 and page_cer <= 0.12 and cover_ok:
        silver = True
    if lines_clean and cluster_bag >= 0.92 and cluster_cer <= 0.12:
        silver = True
    if not silver:
        return "needs_hitl"
    if p_err or not p_eval:
        return "weak_b"
    return "silver"


def to_task_b2_row(
    *,
    image: str,
    ground_truth: str,
    side_matter: str,
    gemini: list[dict[str, Any]],
    post_link: str = "",
    label: str = "",
) -> dict[str, Any]:
    """Task.xlsx B2 row; label = Facebook caption, never GLM.

    Dòng Task.xlsx B2; label = caption Facebook, không phải GLM.
    """
    boxes = []
    for item in gemini:
        boxes.append(
            {
                "text": str(item.get("text") or ""),
                "bounding_box": bbox_of(item),
                "kind": str(item.get("kind") or "ink_text"),
            }
        )
    return {
        "image": image,
        "label": str(label or ""),
        "ground_truth": str(ground_truth or ""),
        "side_matter": str(side_matter or ""),
        "gemini": boxes,
        "post_link": str(post_link or ""),
    }


def _parse_gemini_payload(raw: str) -> tuple[list[dict[str, Any]], float, str | None]:
    """Parse Gemini JSON into B2 boxes + page conf / Parse JSON Gemini thành box B2 + conf trang."""
    parsed = extract_json_object(raw) or {}
    items = parsed.get("gemini") if isinstance(parsed.get("gemini"), list) else []
    boxes: list[dict[str, Any]] = []
    for item in items:
        if not isinstance(item, dict):
            continue
        kind = str(item.get("kind") or "ink_text").strip().lower().replace("-", "_")
        if kind in {"ink", "text", "calligraphy", "body", "main"}:
            kind = "ink_text"
        text = str(item.get("text") or "").strip()
        if not text and kind == "seal":
            text = "[seal]"
        bb = bbox_of(item)
        conf = _clamp_conf(item.get("confidence"), 0.0)
        if not text:
            continue
        boxes.append({"text": text, "bounding_box": bb, "kind": kind, "confidence": conf})
    page_conf = _clamp_conf(parsed.get("confidence"), 0.0)
    if not boxes:
        page_conf = min(page_conf, 0.3) if page_conf else 0.0
        return [], page_conf, None
    if page_conf <= 0:
        vals = [float(b["confidence"]) for b in boxes]
        page_conf = _clamp_conf(sum(vals) / len(vals) if vals else 0.0)
    return boxes, page_conf, None


def _image_size(image_bytes: bytes) -> tuple[int, int]:
    """Return (width, height) of image bytes / Trả (rộng, cao) từ bytes ảnh."""
    try:
        import cv2
        import numpy as np

        arr = np.frombuffer(image_bytes, dtype=np.uint8)
        img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
        if img is not None:
            h, w = img.shape[:2]
            return int(w), int(h)
    except Exception:
        pass
    try:
        from PIL import Image

        img = Image.open(BytesIO(image_bytes))
        return int(img.size[0]), int(img.size[1])
    except Exception:
        return 0, 0


def _downscale_max_side(
    image_bytes: bytes, max_side: int = PADDLE_MAX_SIDE
) -> tuple[bytes, int, int]:
    """Resize like in-cluster Paddle before OCR / Resize giống Paddle trước OCR."""
    width, height = _image_size(image_bytes)
    if width <= 0 or height <= 0:
        return image_bytes, width, height
    longest = max(width, height)
    if longest <= max_side:
        return image_bytes, width, height
    try:
        import cv2
        import numpy as np

        arr = np.frombuffer(image_bytes, dtype=np.uint8)
        img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
        if img is None:
            return image_bytes, width, height
        scale = max_side / float(longest)
        new_w, new_h = max(1, int(width * scale)), max(1, int(height * scale))
        resized = cv2.resize(img, (new_w, new_h), interpolation=cv2.INTER_AREA)
        ok, buf = cv2.imencode(".jpg", resized, [int(cv2.IMWRITE_JPEG_QUALITY), 95])
        if not ok:
            return image_bytes, width, height
        return bytes(buf), new_w, new_h
    except Exception:
        return image_bytes, width, height


def _paddle_sem() -> threading.Semaphore:
    """Process-wide Paddle inflight cap (= replica count) / Trần inflight Paddle cả process."""
    global _PADDLE_SEM
    with _PADDLE_SEM_LOCK:
        if _PADDLE_SEM is None:
            n = max(1, _int_env("FEN_PADDLE_MAX_INFLIGHT", PADDLE_MAX_INFLIGHT))
            _PADDLE_SEM = threading.Semaphore(n)
        return _PADDLE_SEM


def _paddle_ocr(image_bytes: bytes, url: str) -> tuple[dict[str, Any], int, int]:
    """Call in-cluster PaddleOCR; cap inflight to 2 GPUs.

    Gọi PaddleOCR trong cluster; trần inflight = 2 GPU.
    """
    import requests

    payload = {"image_base64": base64.standard_b64encode(image_bytes).decode("ascii")}
    timeout = _int_env("FEN_PADDLE_TIMEOUT_SEC", PADDLE_TIMEOUT_SEC)
    sem = _paddle_sem()
    t_wait = time.time()
    sem.acquire()
    wait_ms = int((time.time() - t_wait) * 1000)
    t_inf = time.time()
    try:
        resp = requests.post(url, json=payload, timeout=timeout)
        resp.raise_for_status()
        data = resp.json() if resp.content else {}
        body = data if isinstance(data, dict) else {}
        infer_ms = int((time.time() - t_inf) * 1000)
        print(f"{LOG} paddle wait_ms={wait_ms} infer_ms={infer_ms}", flush=True)
        return body, wait_ms, infer_ms
    finally:
        sem.release()


def _paddle_track(
    raw_bytes: bytes, paddle_url: str
) -> tuple[list[dict[str, Any]], float, str | None, int, int]:
    """Paddle boxes in 0-1000 space / Box Paddle trong không gian 0-1000."""
    p_blocks: list[dict[str, Any]] = []
    p_conf = 0.0
    p_err: str | None = None
    wait_ms = infer_ms = 0
    try:
        paddle_bytes, pw, ph = _downscale_max_side(raw_bytes)
        pdata, wait_ms, infer_ms = _paddle_ocr(paddle_bytes, paddle_url)
        p_conf = _clamp_conf(pdata.get("confidence"), 0.0)
        for block in pdata.get("blocks") or []:
            if not isinstance(block, dict):
                continue
            bb = poly_to_1000(block.get("box"), pw, ph)
            kind = classify_paddle_block(
                {**block, "bounding_box": bb}, img_w=pw or 1, img_h=ph or 1
            )
            p_blocks.append(
                {
                    "text": to_traditional(str(block.get("text") or "")),
                    "bounding_box": bb,
                    "kind": kind,
                    "confidence": _clamp_conf(block.get("confidence"), 0.0),
                }
            )
    except Exception as exc:
        p_err = str(exc)
        print(f"{LOG} paddle_err err={exc}", flush=True)
    return p_blocks, p_conf, p_err, wait_ms, infer_ms


def _chat_json(prompt: str, model: str) -> tuple[dict[str, Any] | None, float, str | None]:
    """Text-only LLM JSON call / Gọi LLM text-only trả JSON."""
    raw = ""
    try:
        raw, _used = call_chat_completion(prompt, model=model, max_tokens=2048)
    except Exception as exc:
        return None, 0.0, str(exc)
    parsed = extract_json_object(_strip_llm_json(raw or ""))
    if not parsed:
        # One short retry when the first body was empty or non-JSON /
        # Retry ngắn khi lần 1 rỗng hoặc không phải JSON
        try:
            retry_prompt = (
                "Return ONLY a JSON object. First character must be {. No markdown. No <think>.\n"
                + prompt
            )
            raw2, _used2 = call_chat_completion(retry_prompt, model=model, max_tokens=1024)
            parsed = extract_json_object(_strip_llm_json(raw2 or ""))
            raw = raw2 or raw
        except Exception:
            parsed = None
    if not parsed:
        snippet = (raw or "").replace("\n", " ")[:180]
        print(f"{LOG} chat_parse_fail model={model} raw={snippet}", flush=True)
        return None, 0.0, f"parse_failed:{snippet}"
    return parsed, _clamp_conf(parsed.get("confidence"), 0.0), None


def _refine_lines(lines: list[str], model: str) -> tuple[list[str], float, list[str]]:
    """GPT refine; reject invented CJK / GPT tinh chỉnh; từ chối CJK bịa."""
    flags: list[str] = []
    raw_join = "\n".join(lines)
    if not lines:
        return [], 0.0, ["empty_main"]
    parsed, conf, err = _chat_json(GPT_REFINE_PROMPT.format(lines_json=json.dumps(lines, ensure_ascii=False)), model)
    if err or not parsed:
        flags.append("refine_fail")
        return lines, 0.0, flags
    out = parsed.get("calligraphy_lines")
    refined = [str(x).strip() for x in out] if isinstance(out, list) else lines
    refined = [x for x in refined if x]
    if invented_chars(raw_join, "\n".join(refined)):
        flags.append("invented_chars_rejected")
        return lines, 0.0, flags
    layout = str(parsed.get("layout") or "")
    if layout == "unknown" and len(refined) <= 1:
        flags.append("layout_unknown")
    return refined or lines, conf, flags


def _eval_lines(lines: list[str], model: str) -> tuple[list[str], float, str, list[str]]:
    """DeepSeek order eval; reject invented CJK / DeepSeek chấm thứ tự; từ chối CJK bịa."""
    flags: list[str] = []
    raw_join = "\n".join(lines)
    if not lines:
        return [], 0.0, "unknown", ["empty_main"]
    parsed, conf, err = _chat_json(DS_EVAL_PROMPT.format(lines_json=json.dumps(lines, ensure_ascii=False)), model)
    if err or not parsed:
        # Keep GPT lines; do not flag eval_fail on every page /
        # Giữ dòng GPT; không gắn eval_fail lên mọi trang
        print(f"{LOG} eval_skip err={err}", flush=True)
        return lines, 0.0, "unknown", flags
    out = parsed.get("calligraphy_lines")
    refined = [str(x).strip() for x in out] if isinstance(out, list) else lines
    refined = [x for x in refined if x]
    if invented_chars(raw_join, "\n".join(refined)):
        flags.append("invented_chars_rejected")
        return lines, 0.0, "unknown", flags
    order = str(parsed.get("reading_order") or "unknown").strip()
    if order not in {"rtl_columns", "ltr_columns", "grid", "mixed", "unknown"}:
        order = "unknown"
    if order == "unknown" and len(refined) <= 1:
        flags.append("layout_unknown")
    return refined or lines, conf, order, flags


def _eval_or_passthrough(
    g_ref: list[str],
    p_ref: list[str],
    eval_model: str,
) -> tuple[list[str], float, str, list[str], list[str], float, str, list[str]]:
    """DeepSeek permute when model set; else keep GPT refine lines.

    Permute DeepSeek khi có model; không thì giữ dòng sau GPT refine.
    """
    if not (eval_model or "").strip():
        return g_ref, 0.0, "unknown", [], p_ref, 0.0, "unknown", []
    with ThreadPoolExecutor(max_workers=2) as pool:
        fut_da = pool.submit(_eval_lines, g_ref, eval_model)
        fut_db = pool.submit(_eval_lines, p_ref, eval_model) if p_ref else None
        g_eval, ds_a_conf, order_a, f3 = fut_da.result()
        if fut_db:
            p_eval, ds_b_conf, order_b, f4 = fut_db.result()
        else:
            p_eval, ds_b_conf, order_b, f4 = [], 0.0, "unknown", []
    return g_eval, ds_a_conf, order_a, f3, p_eval, ds_b_conf, order_b, f4


def _resolve_eval_model(cfg: Any) -> str:
    """Eval model from env; empty env value disables permute.

    Model eval từ env; giá trị rỗng = tắt permute.
    """
    if "FEN_LABEL_EVAL_MODEL" in os.environ:
        return os.environ.get("FEN_LABEL_EVAL_MODEL", "").strip()
    return str(
        get_value(cfg, "final_exam_nlp", "label_eval_model", fallback=EVAL_MODEL) or ""
    ).strip()


def _weighted_conf(boxes: list[dict[str, Any]]) -> float:
    weights: list[tuple[float, int]] = []
    for box in boxes:
        n = len(compact_cjk(str(box.get("text") or "")))
        if n <= 0:
            continue
        weights.append((_clamp_conf(box.get("confidence"), 0.0), n))
    if not weights:
        return 0.0
    num = sum(c * n for c, n in weights)
    den = sum(n for _c, n in weights)
    return _clamp_conf(num / den if den else 0.0)


def _align_line_lists(lines_a: list[str], lines_b: list[str]) -> list[dict[str, Any]]:
    """Greedy line alignment by fuzzy match / Dóng hàng dòng theo fuzzy."""
    used_b: set[int] = set()
    rows: list[dict[str, Any]] = []
    for ia, ta in enumerate(lines_a):
        best_j, best_s = -1, -1.0
        for ib, tb in enumerate(lines_b):
            if ib in used_b:
                continue
            # Score after S2T so 论/論 and 虽/雖 pair as the same line /
            # Chấm sau S2T để 论/論 và 虽/雖 dóng thành một dòng
            # Bag score pairs scrambled order (春風又綠 vs 綠又風春) /
            # Bag dóng được dòng đảo chữ
            score = max(
                fuzz_ratio(ta, tb),
                partial_ratio(ta, tb),
                bag_dice(ta, tb) * 100.0,
            )
            if score > best_s:
                best_s, best_j = score, ib
        n_a = max(len(compact_cjk(to_traditional(ta))), 1)
        # Short labels (百百 vs 百萬) need a lower pair floor /
        # Nhãn ngắn cần ngưỡng dóng thấp hơn
        pair_floor = 40.0 if n_a <= 2 else 55.0
        if best_j >= 0 and best_s >= pair_floor:
            used_b.add(best_j)
            tb = lines_b[best_j]
            ta_t, tb_t = to_traditional(ta), to_traditional(tb)
            cer_v = cer(ta_t, tb_t)
            ratio = fuzz_ratio(ta_t, tb_t)
            bag = bag_dice(ta_t, tb_t)
            part = partial_ratio(ta_t, tb_t)
            rec = min(len(compact_cjk(ta_t)), len(compact_cjk(tb_t))) / max(
                max(len(compact_cjk(ta_t)), len(compact_cjk(tb_t))), 1
            )
            status = gate_line(cer_v=cer_v, ratio=ratio, bag=bag, partial=part, rec=rec)
            rows.append(
                {
                    "line_id": ia,
                    "text_a": ta,
                    "text_b": tb,
                    "cer": cer_v,
                    "ratio": ratio,
                    "bag": bag,
                    "partial_ratio": part,
                    "line_conf": line_conf_metric(cer_v, ratio),
                    "status": status,
                }
            )
        else:
            rows.append(
                {
                    "line_id": ia,
                    "text_a": ta,
                    "text_b": "",
                    "cer": 1.0,
                    "ratio": 0.0,
                    "bag": 0.0,
                    "partial_ratio": 0.0,
                    "line_conf": 0.0,
                    "status": "a_only",
                }
            )
    for ib, tb in enumerate(lines_b):
        if ib in used_b:
            continue
        rows.append(
            {
                "line_id": f"b{ib}",
                "text_a": "",
                "text_b": tb,
                "cer": 1.0,
                "ratio": 0.0,
                "bag": 0.0,
                "partial_ratio": 0.0,
                "line_conf": 0.0,
                "status": "b_only",
            }
        )
    return rows


def _side_text(boxes: list[dict[str, Any]]) -> str:
    """Join non-calligraphy boxes for side_matter / Ghép box phụ cho side_matter."""
    parts: list[str] = []
    seen: set[str] = set()
    for box in boxes:
        kind = str(box.get("kind") or "")
        if kind in {"ink_text", "main"}:
            continue
        text = str(box.get("text") or "").strip()
        if text and text not in seen:
            seen.add(text)
            parts.append(text)
    return "\n".join(parts)


def _image_rel(image: str) -> str:
    """Relative path under /images/ / Đường dẫn tương đối dưới /images/."""
    rel = str(image or "").split("/images/")[-1].lstrip("/")
    return rel or "image.jpg"


def b2_public_row(row: dict[str, Any]) -> dict[str, Any]:
    """Keep Task.xlsx B2 fields; label is caption only, never GLM.

    Chỉ giữ field B2; label = caption, không copy GLM.
    """
    gemini = row.get("gemini") if isinstance(row.get("gemini"), list) else []
    return {
        "image": str(row.get("image") or ""),
        "label": str(row.get("label") or ""),
        "ground_truth": str(row.get("ground_truth") or ""),
        "side_matter": str(row.get("side_matter") or ""),
        "gemini": gemini,
        "post_link": str(row.get("post_link") or ""),
    }


def b2_submit_row(row: dict[str, Any]) -> dict[str, Any]:
    """Pure Task B2 submit row — no side_matter or extra keys.

    Dòng nộp Task B2 thuần — không side_matter, không cột phụ.
    """
    gemini = row.get("gemini") if isinstance(row.get("gemini"), list) else []
    return {
        "image": str(row.get("image") or ""),
        "label": str(row.get("label") or ""),
        "ground_truth": str(row.get("ground_truth") or ""),
        "gemini": gemini,
        "post_link": str(row.get("post_link") or ""),
    }


def _build_b2_submit_xlsx(rows: list[dict[str, Any]], path: str) -> None:
    """Write pure Task B2 xlsx (no side_matter) / Ghi xlsx B2 thuần (không side_matter)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "B2"
    ws.append(list(B2_SUBMIT_FIELDS))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    link_font = Font(color="0563C1", underline="single")
    for row in rows:
        gemini = row.get("gemini") or []
        ws.append(
            [
                row.get("image") or "",
                row.get("label") or "",
                row.get("ground_truth") or "",
                json.dumps(gemini, ensure_ascii=False),
                row.get("post_link") or "",
            ]
        )
    for excel_row in ws.iter_rows(min_row=2):
        for cell in excel_row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[excel_row[0].row].height = 60
        # Clickable Facebook permalink in column E / Permalink Facebook click được ở cột E
        link_cell = excel_row[4]
        url = str(link_cell.value or "").strip()
        if url.startswith("http"):
            link_cell.hyperlink = url
            link_cell.font = link_font
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 55
    ws.freeze_panes = "A2"
    wb.save(path)


def write_b2_submit_local(
    *,
    group_id: str,
    rows: list[dict[str, Any]],
) -> dict[str, str]:
    """Write pure B2 jsonl+xlsx under /tmp/fen-output/{group_id}/.

    Ghi B2 thuần jsonl+xlsx vào /tmp/fen-output/{group_id}/.
    """
    from common.config import get_output_dir

    submit_rows = [b2_submit_row(r) for r in rows]
    out_dir = get_output_dir() / str(group_id).strip()
    out_dir.mkdir(parents=True, exist_ok=True)
    jsonl_path = out_dir / "task_b2.jsonl"
    xlsx_path = out_dir / "task_b2.xlsx"
    lines = [json.dumps(r, ensure_ascii=False) for r in submit_rows]
    jsonl_path.write_text("\n".join(lines) + ("\n" if lines else ""), encoding="utf-8")
    _build_b2_submit_xlsx(submit_rows, str(xlsx_path))
    print(
        f"{LOG} local_submit n={len(submit_rows)} "
        f"jsonl={jsonl_path} xlsx={xlsx_path}",
        flush=True,
    )
    return {"task_b2_jsonl": str(jsonl_path), "task_b2_xlsx": str(xlsx_path)}


def _copy_src_image(bucket: str, src_key: str, dest_key: str) -> None:
    """Server-side copy into the isolated prefix / Copy phía server vào prefix tách."""
    from minio.commonconfig import CopySource

    if not src_key or not dest_key or src_key == dest_key:
        return
    if not object_exists(bucket, src_key):
        return
    get_minio_client().copy_object(bucket, dest_key, CopySource(bucket, src_key))


def _build_b2_xlsx(rows: list[dict[str, Any]], path: str) -> None:
    """Write Task.xlsx B2 sheet / Ghi sheet B2 đúng Task.xlsx."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "B2"
    ws.append(list(B2_FIELDS))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    link_font = Font(color="0563C1", underline="single")
    for row in rows:
        gemini = row.get("gemini") or []
        ws.append(
            [
                row.get("image") or "",
                row.get("label") or "",
                row.get("ground_truth") or "",
                row.get("side_matter") or "",
                json.dumps(gemini, ensure_ascii=False),
                row.get("post_link") or "",
            ]
        )
    for excel_row in ws.iter_rows(min_row=2):
        for cell in excel_row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[excel_row[0].row].height = 60
        # Clickable Facebook permalink in column F / Permalink Facebook click được ở cột F
        link_cell = excel_row[5]
        url = str(link_cell.value or "").strip()
        if url.startswith("http"):
            link_cell.hyperlink = url
            link_cell.font = link_font
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions["F"].width = 55
    ws.freeze_panes = "A2"
    wb.save(path)


def _excel_cell(value: Any, limit: int = 32000) -> Any:
    """Trim Excel cell to 32k chars / Cắt ô Excel tối đa 32k ký tự."""
    if value is None:
        return ""
    if isinstance(value, (int, float, bool)):
        return value
    text = str(value)
    if len(text) <= limit:
        return text
    return text[: limit - 15] + "\n…[truncated]"


def _excel_num(value: Any, nd: int = 4) -> float | str:
    """Round metric for Excel / Làm tròn số liệu cho Excel."""
    if value is None or value == "":
        return ""
    try:
        return round(float(value), nd)
    except (TypeError, ValueError):
        return ""


def _join_flag_list(flags: Any) -> str:
    """Stable comma-separated flags / Flag cách nhau bởi dấu phẩy, thứ tự ổn định."""
    if not isinstance(flags, list):
        return ""
    return ", ".join(str(x) for x in flags if str(x).strip())


def _line_status_counts(lines: list[dict[str, Any]]) -> str:
    """Count fuse line statuses / Đếm status từng dòng fuse."""
    counts: Counter[str] = Counter()
    for row in lines:
        if isinstance(row, dict):
            counts[str(row.get("status") or "?")] += 1
    return ", ".join(f"{k}={v}" for k, v in sorted(counts.items()))


def _hitl_span_text(spans: Any) -> str:
    """Readable HITL span list / Danh sách span HITL dạng đọc được."""
    lines: list[str] = []
    for span in spans or []:
        if not isinstance(span, dict):
            continue
        reason = str(span.get("reason") or "")
        a = str(span.get("a") or "")
        b = str(span.get("b") or "")
        lines.append(f"{reason}: {a} | {b}".strip(" :|"))
    return "\n".join(lines)


def to_tester_review_row(
    page: dict[str, Any],
    glm: dict[str, Any],
    b2: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """One tester row: OCR + fuse + GLM sidecar, not Task B2.

    Một dòng tester: OCR + fuse + GLM sidecar, không phải Task B2.
    """
    b2 = b2 or {}
    stored = page.get("boxes") if isinstance(page.get("boxes"), dict) else {}
    g_boxes = stored.get("gemini") or []
    p_boxes = stored.get("paddle") or []
    gem = page.get("gemini") if isinstance(page.get("gemini"), dict) else {}
    pad = page.get("paddle") if isinstance(page.get("paddle"), dict) else {}
    gpt_a = page.get("gpt_a") if isinstance(page.get("gpt_a"), dict) else {}
    gpt_b = page.get("gpt_b") if isinstance(page.get("gpt_b"), dict) else {}
    ds_a = page.get("ds_a") if isinstance(page.get("ds_a"), dict) else {}
    ds_b = page.get("ds_b") if isinstance(page.get("ds_b"), dict) else {}
    align = page.get("align") if isinstance(page.get("align"), dict) else {}
    phases = glm.get("phases") if isinstance(glm.get("phases"), dict) else {}
    glm_ph = phases.get("glm") if isinstance(phases.get("glm"), dict) else {}
    metrics = glm.get("metrics") if isinstance(glm.get("metrics"), dict) else {}
    vote = metrics.get("vote") if isinstance(metrics.get("vote"), dict) else {}
    lines = [r for r in (page.get("calligraphy_lines") or []) if isinstance(r, dict)]
    text_a, text_b = _page_ab_text(page)
    if not compact_cjk(text_a):
        text_a = str(glm.get("text_a") or "")
    if not compact_cjk(text_b):
        text_b = str(glm.get("text_b") or "")
    fuse_gt = str(b2.get("ground_truth") or glm.get("fuse_gt") or "")
    side = str(b2.get("side_matter") or _side_text(list(g_boxes) + list(p_boxes)))
    caption = str(page.get("caption") or b2.get("label") or "")
    glm_conf = glm_ph.get("confidence")
    return {
        "page_status": str(page.get("page_status") or glm.get("page_status") or ""),
        "glm_review_bucket": str(glm.get("review_bucket") or glm_ph.get("review_bucket") or ""),
        "glm_agree_with_fuse": bool(metrics.get("agree_with_fuse")),
        "image": str(page.get("image") or glm.get("image") or ""),
        "post_id": str(page.get("post_id") or glm.get("post_id") or ""),
        "quote_batch": page.get("quote_batch") if page.get("quote_batch") is not None else "",
        "post_link": str(page.get("post_link") or glm.get("post_link") or b2.get("post_link") or ""),
        "label_caption": caption,
        "fuse_ground_truth": fuse_gt,
        "glm_recommend": str(glm.get("recommend_ground_truth") or ""),
        "text_a": text_a,
        "text_b": text_b,
        "side_matter": side,
        "flags": _join_flag_list(page.get("flags") or glm.get("flags")),
        "glm_flags": _join_flag_list(glm.get("glm_flags")),
        "hitl_spans": _hitl_span_text(page.get("hitl_spans")),
        "line_status_counts": _line_status_counts(lines),
        "n_lines": len(lines),
        "n_hitl_spans": len(page.get("hitl_spans") or []),
        "n_locked": len(page.get("locked_lines") or []),
        "caption_align": str(align.get("caption_align") or ""),
        "gt_track": str(align.get("gt_track") or ""),
        "page_conf": _excel_num(align.get("page_conf")),
        "bag": _excel_num(align.get("bag")),
        "cluster_bag": _excel_num(align.get("cluster_bag")),
        "compact_bag": _excel_num(align.get("compact_bag")),
        "cer": _excel_num(align.get("cer")),
        "gemini_conf": _excel_num(gem.get("confidence")),
        "paddle_conf": _excel_num(pad.get("confidence")),
        "paddle_error": str(pad.get("error") or ""),
        "n_ink_gemini": int(gem.get("n_ink") or 0),
        "n_ink_paddle": int(pad.get("n_ink") or 0),
        "gpt_a_conf": _excel_num(gpt_a.get("confidence")),
        "gpt_b_conf": _excel_num(gpt_b.get("confidence")),
        "ds_a_order": str(ds_a.get("reading_order") or ""),
        "ds_b_order": str(ds_b.get("reading_order") or ""),
        "ds_a_conf": _excel_num(ds_a.get("confidence")),
        "ds_b_conf": _excel_num(ds_b.get("confidence")),
        "glm_pick": str(glm_ph.get("pick") or ""),
        "glm_source": str(glm_ph.get("recommend_source") or ""),
        "glm_conf": _excel_num(glm_conf),
        "glm_gate_ok": bool(glm_ph.get("gate_ok")),
        "glm_vote_score": _excel_num(glm_ph.get("vote_score") or vote.get("vote_score")),
        "glm_dropped_invented": bool(glm_ph.get("dropped_invented")),
        "glm_retried": bool(glm_ph.get("retried")),
        "rec_vs_fuse_bag": _excel_num(metrics.get("rec_vs_fuse_bag")),
        "rec_vs_fuse_cer": _excel_num(metrics.get("rec_vs_fuse_cer")),
        "rec_vs_a_bag": _excel_num(metrics.get("rec_vs_a_bag")),
        "rec_vs_b_bag": _excel_num(metrics.get("rec_vs_b_bag")),
        "gemini_ink": _ink_lines_from_boxes(g_boxes),
        "paddle_ink": _ink_lines_from_boxes(p_boxes),
        "gemini_boxes_json": json.dumps(g_boxes, ensure_ascii=False),
        "paddle_boxes_json": json.dumps(p_boxes, ensure_ascii=False),
        "seals_json": json.dumps(page.get("seals") or [], ensure_ascii=False),
        "schema_version": str(page.get("schema_version") or ""),
        "glm_schema": str(glm.get("schema_version") or ""),
        "glm_run_id": str(glm.get("run_id") or ""),
        "page_at": str(page.get("at") or ""),
        "glm_at": str(glm.get("at") or ""),
    }


def to_tester_hitl_row(review: dict[str, Any]) -> dict[str, Any]:
    """Compact HITL row; verdict columns stay empty for the tester.

    Dòng HITL gọn; cột kết luận để trống cho tester điền.
    """
    return {
        "verdict": "",
        "corrected_gt": "",
        "notes": "",
        "page_status": str(review.get("page_status") or ""),
        "glm_agree_with_fuse": bool(review.get("glm_agree_with_fuse")),
        "post_link": str(review.get("post_link") or ""),
        "image": str(review.get("image") or ""),
        "post_id": str(review.get("post_id") or ""),
        "label_caption": str(review.get("label_caption") or ""),
        "fuse_ground_truth": str(review.get("fuse_ground_truth") or ""),
        "glm_recommend": str(review.get("glm_recommend") or ""),
        "text_a": str(review.get("text_a") or ""),
        "text_b": str(review.get("text_b") or ""),
        "hitl_spans": str(review.get("hitl_spans") or ""),
        "line_status_counts": str(review.get("line_status_counts") or ""),
        "flags": str(review.get("flags") or ""),
    }


def to_tester_line_rows(
    page: dict[str, Any],
    glm: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    """One Excel row per fused calligraphy line / Mỗi dòng thư pháp một hàng Excel."""
    glm = glm or {}
    image = str(page.get("image") or glm.get("image") or "")
    post_id = str(page.get("post_id") or glm.get("post_id") or "")
    status = str(page.get("page_status") or glm.get("page_status") or "")
    spans = [s for s in (page.get("hitl_spans") or []) if isinstance(s, dict)]
    out: list[dict[str, Any]] = []
    for row in page.get("calligraphy_lines") or []:
        if not isinstance(row, dict):
            continue
        ta = str(row.get("text_a") or "")
        tb = str(row.get("text_b") or "")
        reason = ""
        for span in spans:
            if str(span.get("a") or "") == ta and str(span.get("b") or "") == tb:
                reason = str(span.get("reason") or "")
                break
        out.append(
            {
                "image": image,
                "post_id": post_id,
                "page_status": status,
                "line_id": row.get("line_id") if row.get("line_id") is not None else "",
                "status": str(row.get("status") or ""),
                "fuse": str(row.get("fuse") or ""),
                "text_a": ta,
                "text_b": tb,
                "line_conf": _excel_num(row.get("line_conf")),
                "cer": _excel_num(row.get("cer")),
                "bag": _excel_num(row.get("bag")),
                "ratio": _excel_num(row.get("ratio")),
                "hitl_reason": reason,
            }
        )
    return out


def _append_tester_sheet(
    wb: Any,
    *,
    title: str,
    fields: tuple[str, ...],
    rows: list[dict[str, Any]],
    link_field: str | None = None,
    widths: dict[str, int] | None = None,
) -> None:
    """Write one tester sheet / Ghi một sheet tester."""
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(title)
    ws.append(list(fields))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    link_font = Font(color="0563C1", underline="single")
    link_idx = fields.index(link_field) if link_field and link_field in fields else -1
    for row in rows:
        ws.append([_excel_cell(row.get(name)) for name in fields])
    wrap = Alignment(wrap_text=True, vertical="top")
    for excel_row in ws.iter_rows(min_row=2):
        for cell in excel_row:
            cell.alignment = wrap
        ws.row_dimensions[excel_row[0].row].height = 48
        if link_idx >= 0:
            link_cell = excel_row[link_idx]
            url = str(link_cell.value or "").strip()
            if url.startswith("http"):
                link_cell.hyperlink = url
                link_cell.font = link_font
    for idx, name in enumerate(fields, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = (widths or {}).get(name, 18)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def _append_review_columns_sheet(wb: Any) -> None:
    """Glossary of review (+ hitl fill-in) columns / Từ điển cột review (+ cột điền hitl)."""
    from openpyxl.styles import Alignment, Font

    ws = wb.create_sheet("review_columns")
    headers = ("sheet", "cot", "y_nghia", "tester")
    ws.append(list(headers))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    for name, meaning, tester in TESTER_REVIEW_COLUMN_DOCS:
        ws.append(["review", name, meaning, tester])
    for name, meaning, tester in TESTER_HITL_COLUMN_DOCS:
        ws.append(["hitl", name, meaning, tester])
    for excel_row in ws.iter_rows(min_row=1):
        for cell in excel_row:
            cell.alignment = wrap
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 36
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 78
    ws.column_dimensions["D"].width = 48
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def _build_tester_xlsx(
    *,
    review_rows: list[dict[str, Any]],
    line_rows: list[dict[str, Any]],
    summary: dict[str, Any],
    path: str,
) -> None:
    """Write tester workbook (not Task B2) / Ghi workbook tester (không phải Task B2)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    readme = wb.active
    readme.title = "readme"
    readme_lines = [
        ("Dùng sheet hitl. Không nộp file này làm Task B2.", "Use the hitl sheet. Do not submit this file as Task B2."),
        (
            "So ảnh (post_link) với fuse_ground_truth. Caption ≠ mực trên ảnh.",
            "Compare the photo (post_link) to fuse_ground_truth. Caption is not the ink.",
        ),
        (
            "glm_recommend chỉ gợi ý. Không copy GLM trừ khi khớp nét và chữ nằm trong text_a/text_b.",
            "glm_recommend is a hint. Do not copy GLM unless it matches strokes and A∪B.",
        ),
        (
            "verdict: ok_fuse | use_glm | edit | unreadable | skip. edit thì điền corrected_gt.",
            "verdict: ok_fuse | use_glm | edit | unreadable | skip. If edit, fill corrected_gt.",
        ),
        (
            "Ưu tiên needs_hitl + glm_agree_with_fuse=FALSE. Silver chỉ spot-check.",
            "Prioritize needs_hitl + glm_agree_with_fuse=FALSE. Spot-check silver only.",
        ),
        (
            "label_caption = B2 label. fuse_ground_truth = B2 ground_truth (bản nộp hiện tại).",
            "label_caption = B2 label. fuse_ground_truth = B2 ground_truth (current submit).",
        ),
        (
            "Sheet review_columns = ý nghĩa từng cột review (và 3 cột điền trên hitl).",
            "Sheet review_columns = glossary for review columns (plus hitl fill-in columns).",
        ),
        (f"schema={TESTER_SCHEMA}", f"n_review={summary.get('n_review')} n_lines={summary.get('n_lines')}"),
        (f"exported_at={summary.get('exported_at')}", f"root={summary.get('root')}"),
    ]
    readme.append(["vi", "en"])
    for cell in readme[1]:
        cell.font = Font(bold=True)
    for left, right in readme_lines:
        readme.append([left, right])
    for excel_row in readme.iter_rows(min_row=1):
        for cell in excel_row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    readme.column_dimensions["A"].width = 88
    readme.column_dimensions["B"].width = 88
    readme.row_dimensions[1].height = 22
    for r in range(2, readme.max_row + 1):
        readme.row_dimensions[r].height = 36

    _append_review_columns_sheet(wb)

    summary_ws = wb.create_sheet("summary")
    summary_ws.append(["key", "value"])
    for cell in summary_ws[1]:
        cell.font = Font(bold=True)
    for key, value in summary.items():
        summary_ws.append([key, _excel_cell(value)])
    summary_ws.column_dimensions["A"].width = 28
    summary_ws.column_dimensions["B"].width = 40

    # HITL first, then GLM≠fuse / Ưu tiên needs_hitl rồi GLM khác fuse
    hitl_rows = [to_tester_hitl_row(r) for r in review_rows]
    hitl_rows.sort(
        key=lambda r: (
            0 if r.get("page_status") == "needs_hitl" else 1,
            0 if not r.get("glm_agree_with_fuse") else 1,
            str(r.get("image") or ""),
        )
    )
    _append_tester_sheet(
        wb,
        title="hitl",
        fields=TESTER_HITL_FIELDS,
        rows=hitl_rows,
        link_field="post_link",
        widths={
            "verdict": 14,
            "corrected_gt": 42,
            "notes": 28,
            "page_status": 14,
            "glm_agree_with_fuse": 18,
            "post_link": 48,
            "image": 36,
            "post_id": 22,
            "label_caption": 36,
            "fuse_ground_truth": 48,
            "glm_recommend": 48,
            "text_a": 36,
            "text_b": 36,
            "hitl_spans": 36,
            "line_status_counts": 22,
            "flags": 28,
        },
    )
    # Dropdown for tester verdict / Dropdown kết luận tester
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    hitl_ws = wb["hitl"]
    if hitl_rows:
        verdict_col = get_column_letter(TESTER_HITL_FIELDS.index("verdict") + 1)
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(TESTER_VERDICTS) + '"',
            allow_blank=True,
        )
        dv.error = "Use ok_fuse, use_glm, edit, unreadable, or skip"
        dv.errorTitle = "verdict"
        dv.prompt = "ok_fuse=giữ fuse; use_glm=lấy GLM; edit=sửa corrected_gt"
        dv.promptTitle = "HITL verdict"
        hitl_ws.add_data_validation(dv)
        dv.add(f"{verdict_col}2:{verdict_col}{len(hitl_rows) + 1}")

    review_widths = {
        "image": 36,
        "post_id": 22,
        "post_link": 48,
        "label_caption": 42,
        "fuse_ground_truth": 48,
        "glm_recommend": 48,
        "text_a": 40,
        "text_b": 40,
        "side_matter": 28,
        "flags": 36,
        "glm_flags": 32,
        "hitl_spans": 40,
        "line_status_counts": 28,
        "gemini_ink": 36,
        "paddle_ink": 36,
        "gemini_boxes_json": 28,
        "paddle_boxes_json": 28,
        "seals_json": 24,
        "page_status": 14,
        "glm_review_bucket": 16,
    }
    _append_tester_sheet(
        wb,
        title="review",
        fields=TESTER_REVIEW_FIELDS,
        rows=review_rows,
        link_field="post_link",
        widths=review_widths,
    )
    _append_tester_sheet(
        wb,
        title="lines",
        fields=TESTER_LINE_FIELDS,
        rows=line_rows,
        link_field=None,
        widths={
            "image": 36,
            "post_id": 22,
            "text_a": 40,
            "text_b": 40,
            "page_status": 14,
            "status": 16,
            "fuse": 16,
            "hitl_reason": 16,
        },
    )
    wb.save(path)


def export_tester_workbook(
    *,
    bucket: str,
    root: str,
    local_path: str | None = None,
) -> dict[str, Any]:
    """Excel for testers from complete OCR+fuse+GLM rows; never Task B2.

    Excel cho tester từ ảnh đã OCR+fuse+GLM; không ghi đè Task B2.
    """
    root = _assert_pilot_write_root(root)
    pages_prefix = f"{root}/pages/"
    glm_by = {
        str(r.get("image") or ""): r
        for r in read_jsonl(bucket, f"{root}/glm/recommend.jsonl")
        if r.get("image")
    }
    b2_by = {
        str(r.get("image") or ""): r
        for r in read_jsonl(bucket, f"{root}/task_b2.jsonl")
        if r.get("image")
    }
    keys = [
        k
        for k in list_objects_with_prefix(bucket, pages_prefix, suffix=".json")
        if k.startswith(pages_prefix)
    ]
    review_rows: list[dict[str, Any]] = []
    line_rows: list[dict[str, Any]] = []
    skipped_no_glm = 0
    for key in keys:
        page = _read_json_object(bucket, key) or {}
        image = str(page.get("image") or "")
        glm = glm_by.get(image)
        if not image or not glm:
            skipped_no_glm += 1
            continue
        row = to_tester_review_row(page, glm, b2_by.get(image))
        review_rows.append(row)
        line_rows.extend(to_tester_line_rows(page, glm))
    review_rows.sort(
        key=lambda r: (int(r.get("quote_batch") or 0), str(r.get("image") or ""))
    )
    line_rows.sort(
        key=lambda r: (str(r.get("image") or ""), str(r.get("line_id") or ""))
    )
    n_silver = sum(1 for r in review_rows if r.get("page_status") == "silver")
    n_hitl = sum(1 for r in review_rows if r.get("page_status") == "needs_hitl")
    n_agree = sum(1 for r in review_rows if r.get("glm_agree_with_fuse"))
    n_glm_empty = sum(1 for r in review_rows if not compact_cjk(str(r.get("glm_recommend") or "")))
    summary = {
        "schema": TESTER_SCHEMA,
        "exported_at": utc_now_iso(),
        "root": root,
        "n_pages_listed": len(keys),
        "n_glm": len(glm_by),
        "n_review": len(review_rows),
        "n_lines": len(line_rows),
        "n_skipped_no_glm": skipped_no_glm,
        "n_silver": n_silver,
        "n_needs_hitl": n_hitl,
        "n_glm_agree_fuse": n_agree,
        "n_glm_empty": n_glm_empty,
        "xlsx": f"{root}/tester/review.xlsx",
        "note": "not_task_b2_submission",
    }
    xlsx_key = f"{root}/tester/review.xlsx"
    with tempfile.TemporaryDirectory(prefix="fen-tester-") as tmp:
        local = f"{tmp}/review.xlsx"
        _build_tester_xlsx(
            review_rows=review_rows,
            line_rows=line_rows,
            summary=summary,
            path=local,
        )
        upload_file(get_minio_client(), bucket, xlsx_key, local)
        if local_path:
            # Also keep a laptop copy for testers / Giữ bản local cho tester
            parent = os.path.dirname(os.path.abspath(local_path))
            if parent:
                os.makedirs(parent, exist_ok=True)
            shutil.copy(local, local_path)
            summary["local_path"] = local_path
    upload_json_payload(bucket, f"{root}/tester/manifest.json", summary)
    print(
        f"{LOG} tester xlsx n={summary['n_review']} lines={summary['n_lines']} "
        f"skip_glm={skipped_no_glm} dest={bucket}/{xlsx_key}",
        flush=True,
    )
    return summary


def _gemini_with_retry(image_bytes: bytes, model: str, attempts: int) -> dict[str, Any]:
    """Gemini OCR with retries; do not invent if empty.

    OCR Gemini kèm retry; không bịa chữ khi rỗng.
    """
    history: list[dict[str, Any]] = []
    last: dict[str, Any] = {"boxes": [], "confidence": 0.0, "error": "no_attempt"}
    for i in range(max(1, attempts)):
        processed = preprocess_light(image_bytes) if i == 0 else image_bytes
        raw, err = _vision_ocr(processed, model=model, prompt=GEMINI_PROMPT)
        boxes, conf, _parse_err = _parse_gemini_payload(raw or "")
        ink = [b for b in boxes if b.get("kind") == "ink_text"]
        rec = {
            "attempt": i + 1,
            "confidence": conf,
            "n_boxes": len(boxes),
            "n_ink": len(ink),
            "error": err,
            "conf_source": "model_json",
        }
        history.append(rec)
        last = {"boxes": boxes, "confidence": conf, "error": err, "conf_source": "model_json"}
        if ink:
            break
        time.sleep(1.0)
    last["attempts"] = history
    return last


def _select_sample(
    *,
    bucket: str,
    source_prefix: str,
    group_id: str,
    limit: int,
    done_images: set[str],
) -> list[dict[str, str]]:
    """First N valid posts that still have a MinIO image.

    N bài valid đầu tiên vẫn còn ảnh trên MinIO.
    """
    rows = read_jsonl(bucket, task_export_valid_key(source_prefix, group_id))
    out: list[dict[str, str]] = []
    for row in rows:
        if len(out) >= limit:
            break
        post_id = str(row.get("post_id") or "").strip()
        images = [str(x) for x in (row.get("images") or []) if str(x).strip()]
        for image in images:
            path = image if image.startswith("/") else f"/{image.lstrip('/')}"
            if path in done_images:
                continue
            key = _object_key_from_image_path(source_prefix, group_id, path)
            if not object_exists(bucket, key):
                continue
            out.append(
                {
                    "post_id": post_id,
                    "image": path,
                    "src_key": key,
                    "caption": str(row.get("caption") or row.get("label") or ""),
                    "post_link": facebook_post_link(
                        group_id,
                        post_id,
                        str(row.get("post_link") or row.get("permalink") or ""),
                    ),
                }
            )
            break
    return out


def _caption_map_for_posts(bucket: str, source_prefix: str, group_id: str) -> dict[str, str]:
    """Read-only caption lookup from B1 export / Đọc caption từ export B1, không ghi."""
    out: dict[str, str] = {}
    for row in read_jsonl(bucket, task_export_valid_key(source_prefix, group_id)):
        pid = str(row.get("post_id") or "").strip()
        if not pid:
            continue
        out[pid] = str(row.get("caption") or row.get("label") or "")
    return out


def _post_link_map_for_posts(bucket: str, source_prefix: str, group_id: str) -> dict[str, str]:
    """Read-only Facebook permalinks from B1 export / Permalink Facebook từ export B1."""
    out: dict[str, str] = {}
    for row in read_jsonl(bucket, task_export_valid_key(source_prefix, group_id)):
        pid = str(row.get("post_id") or "").strip()
        if not pid:
            continue
        out[pid] = facebook_post_link(
            group_id,
            pid,
            str(row.get("post_link") or row.get("permalink") or ""),
        )
    return out


def quote_queue_key(root: str, seq: int) -> str:
    """Pilot-prefix queue object for one quote batch / Object hàng đợi một batch quote."""
    root = _assert_pilot_write_root(root)
    return f"{root}/queues/quote_{int(seq):02d}.jsonl"


def priority_longline_queue_key(root: str) -> str:
    """Priority re-OCR queue for long single-line GT / Hàng đợi ưu tiên GT 1 dòng dài."""
    root = _assert_pilot_write_root(root)
    return f"{root}/queues/{PRIORITY_LONGLINE_QUEUE_NAME}"


def _item_page_done(bucket: str, root: str, item: dict[str, Any]) -> bool:
    """True when fuse page sidecar already exists / True khi đã có page fuse."""
    image = str(item.get("image") or "").strip()
    if not image:
        return True
    post_id = str(item.get("post_id") or "").strip() or "unknown"
    return object_exists(bucket, _page_object_key(root, post_id, image))


def _priority_shard_seq_range() -> tuple[int, int]:
    """Inclusive seq range written by 12-way priority shards / Dải seq 12 shard priority."""
    lo = PRIORITY_LONGLINE_SEQ - (QUOTE_BATCH_COUNT - 1)
    return lo, PRIORITY_LONGLINE_SEQ


def _item_priority_reocr_done(bucket: str, root: str, item: dict[str, Any]) -> bool:
    """True when this image was already re-OCR'd by a priority shard.

    True khi ảnh đã được shard priority (seq 88–99) ghi lại — không skip page quote cũ.
    """
    image = str(item.get("image") or "").strip()
    if not image:
        return True
    post_id = str(item.get("post_id") or "").strip() or "unknown"
    page = _read_json_object(bucket, _page_object_key(root, post_id, image)) or {}
    if not page:
        return False
    try:
        qb = int(page.get("quote_batch") or 0)
    except (TypeError, ValueError):
        return False
    lo, hi = _priority_shard_seq_range()
    return lo <= qb <= hi


def select_pending_chunk(
    *,
    bucket: str,
    root: str,
    queue_rows_by_seq: list[tuple[int, list[dict[str, Any]]]],
    target: int,
    force: bool = False,
    is_done: Callable[[str, str, dict[str, Any]], bool] | None = None,
) -> tuple[list[tuple[int, list[dict[str, Any]]]], dict[str, int]]:
    """Pick up to ``target`` pending images; skip done pages unless force.

    Lấy tối đa ``target`` ảnh chưa xong; skip page đã có trừ khi force.
    Round-robin across queues so a chunk does not collapse onto seq 1 /
    Round-robin các queue để chunk không dồn hết vào seq 1.
    """
    pending_by_seq: dict[int, list[dict[str, Any]]] = {}
    skipped_done = 0
    selected = 0
    target_n = max(0, int(target))
    unlimited = target_n <= 0
    done_fn = is_done or _item_page_done

    def _take_pending(seq: int, it: Iterator[dict[str, Any]]) -> bool:
        """Take one not-done row from this queue / Lấy 1 ảnh chưa xong từ queue này."""
        nonlocal skipped_done, selected
        for row in it:
            if not str(row.get("image") or "").strip():
                continue
            if not force and done_fn(bucket, root, row):
                skipped_done += 1
                continue
            pending_by_seq.setdefault(seq, []).append(row)
            selected += 1
            return True
        return False

    seq_iters: list[tuple[int, Iterator[dict[str, Any]]]] = [
        (int(seq), iter(rows)) for seq, rows in queue_rows_by_seq
    ]
    while seq_iters:
        still: list[tuple[int, Iterator[dict[str, Any]]]] = []
        for seq, it in seq_iters:
            if not unlimited and selected >= target_n:
                jobs = [
                    (s, pending_by_seq[s])
                    for s in sorted(pending_by_seq)
                    if pending_by_seq[s]
                ]
                return jobs, {
                    "selected": selected,
                    "skipped_done": skipped_done,
                    "target": target_n,
                    "chunk_full": 1,
                }
            if _take_pending(seq, it):
                still.append((seq, it))
        seq_iters = still

    jobs = [(s, pending_by_seq[s]) for s in sorted(pending_by_seq) if pending_by_seq[s]]
    return jobs, {
        "selected": selected,
        "skipped_done": skipped_done,
        "target": target_n,
        "chunk_full": 0,
    }


def shard_items_for_workers(
    items: list[dict[str, str]],
    n_workers: int,
    *,
    base_seq: int = PRIORITY_LONGLINE_SEQ,
) -> list[tuple[int, list[dict[str, str]]]]:
    """Split one queue into worker shards with distinct part-file seqs.

    Chia một queue thành shard worker, mỗi shard seq riêng để không đè part.
    """
    if not items:
        return []
    n = max(1, min(int(n_workers), len(items)))
    buckets: list[list[dict[str, str]]] = [[] for _ in range(n)]
    # Round-robin so slow images do not pile on one worker /
    # Round-robin để ảnh chậm không dồn một worker
    for i, row in enumerate(items):
        buckets[i % n].append(row)
    jobs: list[tuple[int, list[dict[str, str]]]] = []
    for i, shard in enumerate(buckets):
        if not shard:
            continue
        jobs.append((int(base_seq) - i, shard))
    return jobs


def expand_jobs_for_workers(
    jobs: list[tuple[int, list[dict[str, Any]]]],
    n_workers: int,
    *,
    base_seq: int = QUOTE_PARALLEL_BASE_SEQ,
) -> list[tuple[int, list[dict[str, Any]]]]:
    """Re-shard when chunk has fewer queues than workers.

    Chia lại khi chunk ít queue hơn số worker.
    """
    n_items = sum(len(rows) for _s, rows in jobs)
    if n_workers <= 1 or n_items <= 1 or len(jobs) >= min(int(n_workers), n_items):
        return jobs
    flat: list[dict[str, str]] = []
    for _seq, rows in jobs:
        for row in rows:
            flat.append(row)  # type: ignore[arg-type]
    return shard_items_for_workers(flat, n_workers, base_seq=base_seq)


def split_quote_rows(rows: list[dict[str, Any]], n_batches: int) -> list[list[dict[str, Any]]]:
    """Split rows into contiguous batches; last rem batches get +1.

    Chia hàng liền mạch; rem batch cuối nhận thêm 1 dòng.
    """
    n_batches = max(1, int(n_batches))
    if not rows:
        return [[] for _ in range(n_batches)]
    base = len(rows) // n_batches
    rem = len(rows) % n_batches
    out: list[list[dict[str, Any]]] = []
    idx = 0
    for b in range(n_batches):
        extra = 1 if b >= n_batches - rem else 0
        size = base + extra
        out.append(list(rows[idx : idx + size]))
        idx += size
    return out


def ocr_cjk_from_result(row: dict[str, Any]) -> str:
    """CJK bag from stored OCR result (read-only) / Túi CJK từ ocr_result (chỉ đọc)."""
    parts = [str(row.get("ground_truth") or "")]
    for box in row.get("gemini") or []:
        if isinstance(box, dict):
            parts.append(str(box.get("text") or ""))
    return compact_cjk(to_traditional("".join(parts)))


def upsert_rows_by_image(
    existing: list[dict[str, Any]], incoming: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    """Replace-or-insert rows keyed by image / Thay hoặc thêm dòng theo image."""
    merged: dict[str, dict[str, Any]] = {}
    for row in existing:
        img = str(row.get("image") or "")
        if img:
            merged[img] = row
    for row in incoming:
        img = str(row.get("image") or "")
        if img:
            merged[img] = row
    return list(merged.values())


def collect_quote_items(
    *,
    bucket: str,
    source_prefix: str,
    group_id: str,
) -> list[dict[str, str]]:
    """Build label-dual queue from valid_post images (crawl → OCR).

    Tạo hàng đợi label dual từ ảnh valid_post (crawl → OCR).
    Optional legacy filter: FEN_LABEL_QUOTE_FILTER=true reads ocr_result
    and keeps caption∩OCR quote rows only /
    Lọc legacy tùy chọn: FEN_LABEL_QUOTE_FILTER=true đọc ocr_result
    và chỉ giữ dòng quote caption∩OCR.
    """
    captions = _caption_map_for_posts(bucket, source_prefix, group_id)
    links = _post_link_map_for_posts(bucket, source_prefix, group_id)
    items: list[dict[str, str]] = []
    seen: set[str] = set()

    # Legacy re-queue from prior ocr_result / Hàng đợi legacy từ ocr_result cũ
    if _bool_env("FEN_LABEL_QUOTE_FILTER", False):
        ocr_key = _ocr_output_key(source_prefix, group_id)
        for row in read_jsonl(bucket, ocr_key):
            image = str(row.get("image") or "").strip()
            if not image:
                continue
            path = image if image.startswith("/") else f"/{image.lstrip('/')}"
            if path in seen:
                continue
            post_id = str(row.get("post_id") or "").strip()
            caption = captions.get(post_id) or str(row.get("label") or "")
            ocr_cjk = ocr_cjk_from_result(row)
            align, _hint = caption_use(caption, set(ocr_cjk))
            if align != "quote":
                continue
            src_key = _object_key_from_image_path(source_prefix, group_id, path)
            if not object_exists(bucket, src_key):
                continue
            seen.add(path)
            items.append(
                {
                    "post_id": post_id,
                    "image": path,
                    "src_key": src_key,
                    "caption": caption,
                    "group_id": group_id,
                    "post_link": facebook_post_link(
                        group_id,
                        post_id,
                        str(row.get("post_link") or row.get("permalink") or links.get(post_id) or ""),
                    ),
                }
            )
        return items

    # Exam path: every valid_post image still on MinIO /
    # Luồng exam: mọi ảnh valid_post còn trên MinIO
    for row in read_jsonl(bucket, task_export_valid_key(source_prefix, group_id)):
        post_id = str(row.get("post_id") or "").strip()
        images = [str(x) for x in (row.get("images") or []) if str(x).strip()]
        if not post_id or not images:
            continue
        caption = captions.get(post_id) or str(row.get("caption") or row.get("label") or "")
        post_link = facebook_post_link(
            group_id,
            post_id,
            str(row.get("post_link") or row.get("permalink") or links.get(post_id) or ""),
        )
        for image in images:
            path = image if image.startswith("/") else f"/{image.lstrip('/')}"
            if path in seen:
                continue
            src_key = _object_key_from_image_path(source_prefix, group_id, path)
            if not object_exists(bucket, src_key):
                continue
            seen.add(path)
            items.append(
                {
                    "post_id": post_id,
                    "image": path,
                    "src_key": src_key,
                    "caption": caption,
                    "group_id": group_id,
                    "post_link": post_link,
                }
            )
    return items


def prepare_quote_queues(
    *,
    bucket: str,
    root: str,
    source_prefix: str,
    group_id: str,
    n_batches: int = QUOTE_BATCH_COUNT,
    force: bool = False,
) -> dict[str, Any]:
    """Write quote_{01..N}.jsonl under the pilot prefix / Ghi queue quote dưới prefix pilot.

    Rebuild when force, missing shards, or valid_post image set grew/changed
    (avoids stale queues skipping new crawl images) /
    Rebuild khi force, thiếu shard, hoặc tập ảnh valid_post đổi
    (tránh queue cũ làm OCR skip ảnh crawl mới).
    """
    root = _assert_pilot_write_root(root)
    keys = [quote_queue_key(root, seq) for seq in range(1, n_batches + 1)]
    # Fresh universe from crawl export / Universe mới từ export crawl
    items = collect_quote_items(
        bucket=bucket, source_prefix=source_prefix, group_id=group_id
    )
    wanted = {
        str(r.get("image") or "").strip()
        for r in items
        if str(r.get("image") or "").strip()
    }
    queues_complete = all(object_exists(bucket, k) for k in keys)
    if not force and queues_complete:
        existing_rows: list[dict[str, Any]] = []
        for k in keys:
            existing_rows.extend(read_jsonl(bucket, k))
        have = {
            str(r.get("image") or "").strip()
            for r in existing_rows
            if str(r.get("image") or "").strip()
        }
        # Same image set → keep shards (skip rewrite) /
        # Cùng tập ảnh → giữ shard (không ghi lại)
        if have == wanted:
            sizes = [len(read_jsonl(bucket, k)) for k in keys]
            print(
                f"{LOG} quote queues exist n={sum(sizes)} batches={n_batches} (in sync)",
                flush=True,
            )
            return {
                "prepared": False,
                "n_quote": sum(sizes),
                "batch_sizes": sizes,
                "keys": keys,
            }
        print(
            f"{LOG} quote queues stale have={len(have)} want={len(wanted)} "
            f"added={len(wanted - have)} removed={len(have - wanted)} → rebuild",
            flush=True,
        )
    chunks = split_quote_rows(items, n_batches)
    for seq, chunk in enumerate(chunks, start=1):
        write_jsonl(bucket, quote_queue_key(root, seq), chunk)
    sizes = [len(c) for c in chunks]
    print(
        f"{LOG} prepared quote queues n={len(items)} batches={n_batches} sizes={sizes}",
        flush=True,
    )
    return {"prepared": True, "n_quote": len(items), "batch_sizes": sizes, "keys": keys}


def replay_existing_pilot_pages(
    *,
    bucket: str,
    root: str,
    source_prefix: str,
    group_id: str,
) -> dict[str, Any]:
    """Re-fuse saved pages under label_dual_pilot only; no Gemini/Paddle.

    Fuse lại page đã OCR chỉ trong label_dual_pilot; không gọi Gemini/Paddle.
    """
    root = _assert_pilot_write_root(root)
    pages_prefix = f"{root}/pages/"
    b2_key = f"{root}/task_b2.jsonl"
    captions = _caption_map_for_posts(bucket, source_prefix, group_id)
    keys = [
        k
        for k in list_objects_with_prefix(bucket, pages_prefix, suffix=".json")
        if k.startswith(pages_prefix)
    ]
    print(f"{LOG} replay pages={len(keys)} dest={bucket}/{root}", flush=True)
    existing_b2 = {
        str(r.get("image") or ""): b2_public_row(r)
        for r in read_jsonl(bucket, b2_key)
        if r.get("image")
    }
    flags_all: list[dict[str, Any]] = []
    n_silver = 0
    for key in keys:
        page = _read_json_object(bucket, key) or {}
        if not page.get("calligraphy_lines"):
            continue
        post_id = str(page.get("post_id") or "")
        image = str(page.get("image") or "")
        post_link = facebook_post_link(
            group_id, post_id, str(page.get("post_link") or "")
        )
        rec = recompute_from_page(page, caption=captions.get(post_id, ""))
        page["schema_version"] = SCHEMA
        page["post_link"] = post_link
        page["caption"] = captions.get(post_id) or str(page.get("caption") or "")
        page["flags"] = rec.get("flags") or collect_page_flags(
            page.get("flags") or [],
            aligned=rec["aligned"],
            fused={"hitl_spans": rec["hitl_spans"]},
            wrap_recovered=False,
            g_eval=[str(r.get("text_a") or "") for r in rec["aligned"]],
            p_eval=[str(r.get("text_b") or "") for r in rec["aligned"]],
        )
        page["page_status"] = rec["page_status"]
        page["calligraphy_lines"] = rec["aligned"]
        page["hitl_spans"] = rec["hitl_spans"]
        page["locked_lines"] = rec["locked_lines"]
        align = page.get("align") if isinstance(page.get("align"), dict) else {}
        align["gt_track"] = rec["gt_track"]
        align["compact_bag"] = rec["compact_bag"]
        align["caption_align"] = rec["caption_align"]
        page["align"] = align
        page["at"] = utc_now_iso()
        upload_json_payload(bucket, key, page)
        prev = existing_b2.get(image) or to_task_b2_row(
            image=image, ground_truth="", side_matter="", gemini=[]
        )
        existing_b2[image] = b2_public_row(
            {
                **prev,
                "ground_truth": rec["ground_truth"],
                "label": captions.get(post_id) or str(prev.get("label") or ""),
                "post_link": post_link,
            }
        )
        flags_all.append(
            {
                "image": image,
                "post_id": post_id,
                "post_link": post_link,
                "flags": list(page.get("flags") or []),
                "status": rec["page_status"],
            }
        )
        if rec["page_status"] == "silver":
            n_silver += 1
        print(
            f"{LOG} replay {post_id} {rec['old_status']}→{rec['page_status']} "
            f"cap={rec['caption_align']}",
            flush=True,
        )
    public_rows = [b2_public_row(r) for r in existing_b2.values()]
    write_jsonl(bucket, b2_key, public_rows)
    with tempfile.TemporaryDirectory(prefix="fen-b2-") as tmp:
        xlsx = f"{tmp}/task_b2.xlsx"
        _build_b2_xlsx(public_rows, xlsx)
        upload_file(get_minio_client(), bucket, f"{root}/task_b2.xlsx", xlsx)
    write_jsonl(bucket, f"{root}/flags.jsonl", flags_all)
    summary = {
        "schema_version": SCHEMA,
        "n_images": len(public_rows),
        "processed_this_run": len(flags_all),
        "n_silver": n_silver,
        "n_flagged": len(flags_all) - n_silver,
        "n_silver_this_run": n_silver,
        "force": False,
        "replay": True,
        "task_b2": b2_key,
        "task_b2_xlsx": f"{root}/task_b2.xlsx",
        "images_prefix": f"{root}/images/",
        "root": root,
        "updated_at": utc_now_iso(),
    }
    upload_json_payload(bucket, f"{root}/summary.json", summary)
    upload_json_payload(
        bucket,
        f"{root}/checkpoint.json",
        {
            "schema_version": SCHEMA,
            "replay": True,
            "n_b2": len(public_rows),
            "last_flush_at": utc_now_iso(),
        },
    )
    print(f"{LOG} replay done {summary}", flush=True)
    return summary


def _ink_lines_from_boxes(boxes: Any) -> str:
    """Join ink_text box strings for GLM A/B / Ghép chữ box mực cho GLM A/B."""
    lines: list[str] = []
    for box in boxes or []:
        if not isinstance(box, dict):
            continue
        if str(box.get("kind") or "ink_text") != "ink_text":
            continue
        text = str(box.get("text") or "").strip()
        if compact_cjk(text):
            lines.append(text)
    return "\n".join(lines)


def _page_ab_text(page: dict[str, Any]) -> tuple[str, str]:
    """Join A/B verses from page lines / Ghép câu A/B từ dòng page."""
    a_lines: list[str] = []
    b_lines: list[str] = []
    for row in page.get("calligraphy_lines") or []:
        ta = str(row.get("text_a") or "").strip()
        tb = str(row.get("text_b") or "").strip()
        if ta:
            a_lines.append(ta)
        if tb:
            b_lines.append(tb)
    a_text = "\n".join(a_lines)
    b_text = "\n".join(b_lines)
    # Refine may collapse verses to "..." with no CJK; use ink boxes /
    # Refine có thể gom câu thành "..." không CJK; dùng box mực
    if not compact_cjk(a_text) and not compact_cjk(b_text):
        stored = page.get("boxes") if isinstance(page.get("boxes"), dict) else {}
        a_text = _ink_lines_from_boxes(stored.get("gemini") or [])
        b_text = _ink_lines_from_boxes(stored.get("paddle") or [])
    return a_text, b_text


def _boxes_with_conf(boxes: Any) -> list[dict[str, Any]]:
    """Copy boxes keeping confidence / Copy box giữ confidence."""
    out: list[dict[str, Any]] = []
    for box in boxes or []:
        if not isinstance(box, dict):
            continue
        out.append(
            {
                "text": str(box.get("text") or ""),
                "bounding_box": bbox_of(box),
                "kind": str(box.get("kind") or "ink_text"),
                "confidence": _clamp_conf(box.get("confidence"), 0.0),
            }
        )
    return out


def _mean_or_none(vals: list[float]) -> float | None:
    """Mean of non-empty list / Trung bình list khác rỗng."""
    return round(sum(vals) / len(vals), 4) if vals else None


def collect_glm_flags(
    *,
    page_status: str,
    fuse_gt: str,
    gate_ok: bool,
    dropped_invented: bool,
    retried: bool,
    source: str,
    pick: str,
    agree: bool,
    glm_conf: float,
    pick_mismatch: bool = False,
    review_bucket: str = "",
    mix: bool = False,
) -> list[str]:
    """Flags for GLM log row only / Flag chỉ trên dòng log GLM."""
    out: list[str] = []
    if dropped_invented:
        out.append("glm_dropped_invented")
    if retried:
        out.append("glm_retry")
    if str(source).startswith("fallback"):
        out.append("glm_fallback")
    if source == "glm_empty_ab":
        out.append("glm_empty_ab")
    if not gate_ok:
        out.append("glm_gate_fail")
    else:
        out.append("glm_gate_ok")
    if pick == "parse_fail":
        out.append("glm_parse_fail")
    if pick == "neither":
        out.append("glm_pick_neither")
    if pick == "b":
        out.append("glm_pick_b")
    if pick == "a":
        out.append("glm_pick_a")
    if not compact_cjk(fuse_gt):
        out.append("glm_fuse_empty")
    elif page_status == "silver" and agree:
        out.append("glm_agree_fuse")
    elif page_status == "silver" and not agree:
        out.append("glm_disagree_fuse")
    if glm_conf < 0.5:
        out.append("glm_low_conf")
    if pick_mismatch:
        out.append("glm_pick_mismatch")
    if mix:
        out.append("glm_mix")
    if review_bucket:
        out.append(f"glm_{review_bucket}")
    return sorted(set(out))


def build_glm_log_row(
    *,
    page: dict[str, Any],
    b2: dict[str, Any] | None,
    recommend: str,
    pick: str,
    source: str,
    glm_conf: float,
    retried: bool,
    dropped_invented: bool,
    latency_ms: int,
    run_id: str,
    model: str,
) -> dict[str, Any]:
    """Assemble one glm/recommend.jsonl object / Lắp một object log GLM."""
    text_a, text_b = _page_ab_text(page)
    fuse_gt = str((b2 or {}).get("ground_truth") or "")
    rec = accept_ab_recommend(text_a, text_b, recommend)
    gate_ok = bool(rec)
    agree = bool(compact_cjk(fuse_gt) and norm_cjk(rec) == norm_cjk(fuse_gt))
    vote = compute_glm_vote(
        text_a=text_a,
        text_b=text_b,
        recommend=recommend,
        pick=pick,
        source=source,
        glm_conf=glm_conf,
        gate_ok=gate_ok,
        dropped_invented=dropped_invented,
    )
    stored = page.get("boxes") if isinstance(page.get("boxes"), dict) else {}
    g_boxes = _boxes_with_conf(stored.get("gemini") or (b2 or {}).get("gemini") or [])
    p_boxes = _boxes_with_conf(stored.get("paddle") or [])
    g_ink = [b for b in g_boxes if b.get("kind") == "ink_text"]
    p_ink = [b for b in p_boxes if b.get("kind") == "ink_text"]
    gem = page.get("gemini") if isinstance(page.get("gemini"), dict) else {}
    pad = page.get("paddle") if isinstance(page.get("paddle"), dict) else {}
    gpt_a = page.get("gpt_a") if isinstance(page.get("gpt_a"), dict) else {}
    gpt_b = page.get("gpt_b") if isinstance(page.get("gpt_b"), dict) else {}
    ds_a = page.get("ds_a") if isinstance(page.get("ds_a"), dict) else {}
    ds_b = page.get("ds_b") if isinstance(page.get("ds_b"), dict) else {}
    align = page.get("align") if isinstance(page.get("align"), dict) else {}
    lines = list(page.get("calligraphy_lines") or [])
    g_confs = [float(b["confidence"]) for b in g_boxes if b.get("confidence")]
    p_confs = [float(b["confidence"]) for b in p_boxes if b.get("confidence")]
    g_ink_c = [float(b["confidence"]) for b in g_ink if b.get("confidence")]
    p_ink_c = [float(b["confidence"]) for b in p_ink if b.get("confidence")]
    page_status = str(page.get("page_status") or "")
    glm_flags = collect_glm_flags(
        page_status=page_status,
        fuse_gt=fuse_gt,
        gate_ok=gate_ok,
        dropped_invented=dropped_invented,
        retried=retried,
        source=source,
        pick=pick,
        agree=agree,
        glm_conf=glm_conf,
        pick_mismatch=bool(vote.get("pick_mismatch")),
        review_bucket=str(vote.get("review_bucket") or ""),
        mix=float(vote.get("mix_rate") or 0) > 0,
    )
    rec_vs_fuse_bag = bag_dice(rec, fuse_gt) if compact_cjk(fuse_gt) else None
    rec_vs_fuse_cer = cer(fuse_gt, rec) if compact_cjk(fuse_gt) else None
    return {
        "schema_version": GLM_LOG_SCHEMA,
        "run_id": run_id,
        "at": utc_now_iso(),
        "image": str(page.get("image") or ""),
        "post_id": str(page.get("post_id") or ""),
        "post_link": str(page.get("post_link") or (b2 or {}).get("post_link") or ""),
        "page_status": page_status,
        "flags": list(page.get("flags") or []),
        "glm_flags": glm_flags,
        "text_a": text_a,
        "text_b": text_b,
        "fuse_gt": fuse_gt,
        "recommend_ground_truth": rec,
        "phases": {
            "gemini": {
                "confidence": _clamp_conf(gem.get("confidence"), 0.0),
                "conf_source": str(gem.get("conf_source") or "model_json"),
                "n_boxes": len(g_boxes),
                "n_ink": int(gem.get("n_ink") or len(g_ink)),
                "attempts": len(gem.get("attempts") or []) or None,
            },
            "paddle": {
                "confidence": _clamp_conf(pad.get("confidence"), 0.0),
                "conf_source": str(pad.get("conf_source") or "detector"),
                "n_boxes": len(p_boxes),
                "n_ink": int(pad.get("n_ink") or len(p_ink)),
                "error": pad.get("error"),
                "wait_ms": pad.get("wait_ms"),
                "infer_ms": pad.get("infer_ms"),
            },
            "gpt_a": {
                "confidence": _clamp_conf(gpt_a.get("confidence"), 0.0),
                "conf_source": "model_json",
            },
            "gpt_b": {
                "confidence": _clamp_conf(gpt_b.get("confidence"), 0.0),
                "conf_source": "model_json",
            },
            "ds_a": {
                "confidence": _clamp_conf(ds_a.get("confidence"), 0.0),
                "conf_source": "model_json",
                "reading_order": ds_a.get("reading_order"),
            },
            "ds_b": {
                "confidence": _clamp_conf(ds_b.get("confidence"), 0.0),
                "conf_source": "model_json",
                "reading_order": ds_b.get("reading_order"),
            },
            "align": {
                "confidence": _clamp_conf(align.get("page_conf"), 0.0),
                "conf_source": "metric",
                "cer": align.get("cer"),
                "bag": align.get("bag"),
                "cluster_bag": align.get("cluster_bag"),
                "cluster_cer": align.get("cluster_cer"),
                "compact_bag": align.get("compact_bag"),
                "caption_align": align.get("caption_align"),
            },
            "fuse": {
                "gt_track": align.get("gt_track"),
                "n_locked": len(page.get("locked_lines") or []),
                "n_hitl_spans": len(page.get("hitl_spans") or []),
            },
            "glm": {
                "confidence": _clamp_conf(glm_conf, 0.0),
                "conf_source": "model_json",
                "model": model,
                "pick": pick,
                "recommend_source": source,
                "retried": retried,
                "gate_ok": gate_ok,
                "dropped_invented": dropped_invented,
                "latency_ms": latency_ms,
                "vote_score": vote.get("vote_score"),
                "review_bucket": vote.get("review_bucket"),
            },
        },
        "boxes": {"gemini": g_boxes, "paddle": p_boxes},
        "lines": lines,
        "metrics": {
            "n_verses_a": len([x for x in text_a.splitlines() if compact_cjk(x)]),
            "n_verses_b": len([x for x in text_b.splitlines() if compact_cjk(x)]),
            "n_verses_rec": len([x for x in rec.splitlines() if compact_cjk(x)]),
            "n_cjk_a": len(compact_cjk(text_a)),
            "n_cjk_b": len(compact_cjk(text_b)),
            "n_cjk_rec": len(compact_cjk(rec)),
            "agree_with_fuse": agree,
            "rec_vs_fuse_bag": rec_vs_fuse_bag,
            "rec_vs_fuse_cer": rec_vs_fuse_cer,
            "rec_vs_a_bag": bag_dice(rec, text_a),
            "rec_vs_a_cer": cer(text_a, rec),
            "rec_vs_b_bag": bag_dice(rec, text_b),
            "rec_vs_b_cer": cer(text_b, rec),
            "box_conf_gemini_mean": _mean_or_none(g_confs),
            "box_conf_gemini_min": min(g_confs) if g_confs else None,
            "box_conf_paddle_mean": _mean_or_none(p_confs),
            "box_conf_paddle_min": min(p_confs) if p_confs else None,
            "ink_conf_gemini_mean": _mean_or_none(g_ink_c),
            "ink_conf_paddle_mean": _mean_or_none(p_ink_c),
            "vote": vote,
        },
        "review_bucket": vote.get("review_bucket"),
    }


def _glm_msg_text(msg: Any) -> str:
    """Content plus GLM reasoning / Content cộng reasoning GLM."""
    text = (getattr(msg, "content", None) or "").strip()
    if text:
        return text
    return str(
        getattr(msg, "reasoning_content", None) or getattr(msg, "reasoning", None) or ""
    ).strip()


def _glm_vision_raw(image_bytes: bytes, prompt: str, model: str) -> tuple[str, str | None]:
    """Vision call that also reads GLM thinking / Gọi vision, đọc cả thinking GLM."""
    try:
        from openai import OpenAI
    except ModuleNotFoundError as exc:
        return "", f"missing_openai:{exc}"
    from common.fen_stage_config import stage_api_key, stage_base_url

    cfg = load_config()
    try:
        from common.api_keys import next_api_key

        api_key = next_api_key()
    except Exception:
        # Prefer fen_label_glm, then gemini_opencv / Ưu tiên fen_label_glm rồi gemini_opencv
        api_key = stage_api_key("fen_label_glm") or get_value(
            cfg, "gemini_opencv", "api_key", fallback=""
        ).strip()
    if not api_key:
        return "", "missing_gemini_api_key"
    base_url = (
        stage_base_url("fen_label_glm")
        or get_value(cfg, "gemini_opencv", "base_url", fallback="https://ramclouds.me/v1").strip()
    )
    png = _to_png_bytes(image_bytes, 1600)
    b64 = base64.standard_b64encode(png).decode("ascii")
    client = OpenAI(api_key=api_key, base_url=base_url or None)
    try:
        print(f"{LOG} glm vision model={model}", flush=True)
        response = client.chat.completions.create(
            model=model,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64}"}},
                    ],
                }
            ],
            max_tokens=4096,
            temperature=0.1,
        )
        raw = _glm_msg_text(response.choices[0].message)
        return raw, None if raw else "empty_reply"
    except Exception as exc:
        return "", str(exc)


def _parse_glm_recommend(raw: str, text_a: str, text_b: str) -> tuple[str, str, str, bool, float]:
    """Parse pick, gated recommend, raw, dropped, conf.

    Parse pick, recommend đã cổng, raw, dropped, conf.
    """
    parsed = extract_json_object(_THINK_RE.sub("", raw or "")) or {}
    pick = str(parsed.get("pick") or "").strip().lower()
    if pick not in {"a", "b", "neither"}:
        pick = "parse_fail"
    rec_raw = str(
        parsed.get("recommend_ground_truth")
        or parsed.get("recommend")
        or ""
    ).strip()
    rec = accept_ab_recommend(text_a, text_b, rec_raw)
    dropped = bool(rec_raw) and not rec
    conf = _clamp_conf(parsed.get("confidence"), 0.0)
    return pick, rec, rec_raw, dropped, conf


def fallback_recommend(text_a: str, text_b: str, pick: str) -> tuple[str, str]:
    """Force gated recommend from A or B; never invent.

    Bắt buộc recommend đã cổng từ A hoặc B; không bịa chữ.
    Empty only when A∪B has no CJK / Rỗng chỉ khi A∪B không có CJK.
    """
    a_ok = accept_ab_recommend(text_a, text_b, text_a)
    b_ok = accept_ab_recommend(text_a, text_b, text_b)
    pick_l = str(pick or "").strip().lower()
    if pick_l == "a" and a_ok:
        return a_ok, "fallback_pick_a"
    if pick_l == "b" and b_ok:
        return b_ok, "fallback_pick_b"
    # Prefer the named track, then the other / Ưu tiên nhánh pick, rồi nhánh còn lại
    if pick_l == "a" and b_ok:
        return b_ok, "fallback_pick_a_then_b"
    if pick_l == "b" and a_ok:
        return a_ok, "fallback_pick_b_then_a"
    if a_ok:
        return a_ok, "fallback_page_a"
    if b_ok:
        return b_ok, "fallback_page_b"
    return "", "glm_empty_ab"


def judge_glm_page(image_bytes: bytes, text_a: str, text_b: str, model: str) -> dict[str, Any]:
    """Force gated recommend_ground_truth; not silver.

    Bắt buộc recommend đã cổng; không silver.
    """
    t0 = time.time()
    raw, err = _glm_vision_raw(image_bytes, GLM_PROMPT.format(a=text_a, b=text_b), model)
    pick = "parse_fail"
    rec = ""
    dropped = False
    conf = 0.0
    source = "model"
    retried = False
    if err and not raw:
        rec, source = fallback_recommend(text_a, text_b, "neither")
        pick = "neither" if source == "glm_empty_ab" else (
            "b" if source.endswith("_b") or "then_b" in source else "a"
        )
        return {
            "pick": pick,
            "recommend": rec,
            "source": source if rec or source == "glm_empty_ab" else "error",
            "conf": 0.0,
            "retried": False,
            "dropped": False,
            "raw": err,
            "latency_ms": int((time.time() - t0) * 1000),
        }
    pick, rec, rec_raw, dropped, conf = _parse_glm_recommend(raw, text_a, text_b)
    source = "model"
    if not rec:
        retried = True
        raw2, _err2 = _glm_vision_raw(
            image_bytes, GLM_RETRY_PROMPT.format(a=text_a, b=text_b), model
        )
        pick2, rec2, rec_raw2, dropped2, conf2 = _parse_glm_recommend(raw2, text_a, text_b)
        if rec2:
            pick, rec, rec_raw, dropped, conf, raw = pick2, rec2, rec_raw2, dropped2, conf2, raw2
            source = "retry"
        else:
            dropped = dropped or dropped2
            raw = (raw or "") + "\n---retry---\n" + (raw2 or "")
    if not rec:
        rec, source = fallback_recommend(text_a, text_b, pick)
        if source == "glm_empty_ab":
            pick = "neither"
    return {
        "pick": pick,
        "recommend": rec,
        "source": source,
        "conf": conf,
        "retried": retried,
        "dropped": dropped,
        "raw": (raw or "")[:800],
        "latency_ms": int((time.time() - t0) * 1000),
    }


def _glm_summary_from_rows(rows: list[dict[str, Any]], run_id: str, model: str) -> dict[str, Any]:
    """Aggregate GLM log metrics / Cộng chỉ số log GLM."""
    lat = [int((r.get("phases") or {}).get("glm", {}).get("latency_ms") or 0) for r in rows]
    lat = [x for x in lat if x > 0]
    def _phase_mean(name: str) -> float | None:
        vals = [
            float((r.get("phases") or {}).get(name, {}).get("confidence") or 0)
            for r in rows
            if (r.get("phases") or {}).get(name)
        ]
        vals = [v for v in vals if v > 0]
        return _mean_or_none(vals)

    glm_flags = [f for r in rows for f in (r.get("glm_flags") or [])]
    votes = [
        (r.get("metrics") or {}).get("vote") or {}
        for r in rows
        if (r.get("metrics") or {}).get("vote")
    ]
    scores = [float(v.get("vote_score") or 0) for v in votes]
    return {
        "schema_version": GLM_LOG_SCHEMA,
        "run_id": run_id,
        "model": model,
        "n": len(rows),
        "n_gate_ok": sum(1 for r in rows if "glm_gate_ok" in (r.get("glm_flags") or [])),
        "n_gate_fail": sum(1 for r in rows if "glm_gate_fail" in (r.get("glm_flags") or [])),
        "n_dropped_invented": sum(1 for r in rows if "glm_dropped_invented" in (r.get("glm_flags") or [])),
        "n_retry": sum(1 for r in rows if "glm_retry" in (r.get("glm_flags") or [])),
        "n_fallback": sum(1 for r in rows if "glm_fallback" in (r.get("glm_flags") or [])),
        "n_empty_ab": sum(1 for r in rows if "glm_empty_ab" in (r.get("glm_flags") or [])),
        "n_pick_a": sum(1 for r in rows if "glm_pick_a" in (r.get("glm_flags") or [])),
        "n_pick_b": sum(1 for r in rows if "glm_pick_b" in (r.get("glm_flags") or [])),
        "n_pick_neither": sum(1 for r in rows if "glm_pick_neither" in (r.get("glm_flags") or [])),
        "n_parse_fail": sum(1 for r in rows if "glm_parse_fail" in (r.get("glm_flags") or [])),
        "n_silver": sum(1 for r in rows if r.get("page_status") == "silver"),
        "n_hitl": sum(1 for r in rows if r.get("page_status") == "needs_hitl"),
        "n_agree_silver": sum(1 for r in rows if "glm_agree_fuse" in (r.get("glm_flags") or [])),
        "n_disagree_silver": sum(1 for r in rows if "glm_disagree_fuse" in (r.get("glm_flags") or [])),
        "n_fuse_empty": sum(1 for r in rows if "glm_fuse_empty" in (r.get("glm_flags") or [])),
        "n_mix": sum(1 for r in rows if "glm_mix" in (r.get("glm_flags") or [])),
        "n_pick_mismatch": sum(1 for r in rows if "glm_pick_mismatch" in (r.get("glm_flags") or [])),
        "n_review_ok": sum(1 for r in rows if r.get("review_bucket") == "review_ok"),
        "n_review_mix": sum(1 for r in rows if r.get("review_bucket") == "review_mix"),
        "n_review_weak": sum(1 for r in rows if r.get("review_bucket") == "review_weak"),
        "n_review_empty": sum(1 for r in rows if r.get("review_bucket") == "review_empty"),
        "mean_vote_score": _mean_or_none(scores),
        "mean_n_glyph_pick_a": _mean_or_none([float(v.get("n_pick_a") or 0) for v in votes]),
        "mean_n_glyph_pick_b": _mean_or_none([float(v.get("n_pick_b") or 0) for v in votes]),
        "mean_conf_gemini": _phase_mean("gemini"),
        "mean_conf_paddle": _phase_mean("paddle"),
        "mean_conf_gpt_a": _phase_mean("gpt_a"),
        "mean_conf_gpt_b": _phase_mean("gpt_b"),
        "mean_conf_ds_a": _phase_mean("ds_a"),
        "mean_conf_ds_b": _phase_mean("ds_b"),
        "mean_conf_align": _phase_mean("align"),
        "mean_conf_glm": _phase_mean("glm"),
        "p50_latency_ms": int(statistics.median(lat)) if lat else None,
        "p95_latency_ms": int(sorted(lat)[max(0, int(0.95 * (len(lat) - 1)))]) if lat else None,
        "glm_flag_counts": dict(Counter(glm_flags)),
        "updated_at": utc_now_iso(),
    }


def run_glm_recommend_pass(
    *,
    bucket: str,
    root: str,
    flush_posts: int = 10,
    force: bool = False,
    model: str | None = None,
) -> dict[str, Any]:
    """GLM recommend for every page; write glm/ sidecar only.

    GLM recommend mọi page; chỉ ghi sidecar glm/.
    """
    root = _assert_pilot_write_root(root)
    model = (model or os.environ.get("FEN_LABEL_GLM_MODEL") or GLM_MODEL).strip()
    run_id = os.environ.get("FEN_LABEL_GLM_RUN_ID") or f"glm_rec_{time.strftime('%Y%m%d_%H%M%S')}"
    pages_prefix = f"{root}/pages/"
    rec_key = f"{root}/glm/recommend.jsonl"
    sum_key = f"{root}/glm/summary.json"
    raw_key = f"{root}/glm/runs/{run_id}.jsonl"
    b2_by_image = {
        str(r.get("image") or ""): r
        for r in read_jsonl(bucket, f"{root}/task_b2.jsonl")
        if r.get("image")
    }
    existing = {
        str(r.get("image") or ""): r
        for r in read_jsonl(bucket, rec_key)
        if r.get("image")
    }
    keys = [
        k
        for k in list_objects_with_prefix(bucket, pages_prefix, suffix=".json")
        if k.startswith(pages_prefix)
    ]
    print(f"{LOG} glm pages={len(keys)} skip={0 if force else len(existing)} model={model}", flush=True)
    pending: list[dict[str, Any]] = []
    pending_raw: list[dict[str, Any]] = []
    n_ok = n_fail = 0

    def _flush_glm() -> None:
        if not pending:
            return
        merged = dict(existing)
        for row in pending:
            merged[str(row["image"])] = row
        rows = list(merged.values())
        write_jsonl(bucket, rec_key, rows)
        existing.update({str(r["image"]): r for r in pending})
        prev_raw = read_jsonl(bucket, raw_key)
        write_jsonl(bucket, raw_key, prev_raw + pending_raw)
        glm_sum = _glm_summary_from_rows(list(existing.values()), run_id, model)
        upload_json_payload(bucket, sum_key, glm_sum)
        pending.clear()
        pending_raw.clear()
        print(f"{LOG} glm flush n={len(existing)}", flush=True)

    for i, key in enumerate(keys, start=1):
        page = _read_json_object(bucket, key) or {}
        image = str(page.get("image") or "")
        if not image:
            continue
        if not force and image in existing:
            continue
        text_a, text_b = _page_ab_text(page)
        if not text_a and not text_b:
            n_fail += 1
            continue
        img_key = f"{root}/images/{_image_rel(image)}"
        try:
            image_bytes = _read_object_bytes(bucket, img_key)
        except Exception as exc:
            print(f"{LOG} glm miss image {img_key} {exc}", flush=True)
            n_fail += 1
            continue
        judged = judge_glm_page(image_bytes, text_a, text_b, model)
        row = build_glm_log_row(
            page=page,
            b2=b2_by_image.get(image),
            recommend=str(judged.get("recommend") or ""),
            pick=str(judged.get("pick") or "parse_fail"),
            source=str(judged.get("source") or "model"),
            glm_conf=float(judged.get("conf") or 0.0),
            retried=bool(judged.get("retried")),
            dropped_invented=bool(judged.get("dropped")),
            latency_ms=int(judged.get("latency_ms") or 0),
            run_id=run_id,
            model=model,
        )
        pending.append(row)
        pending_raw.append(
            {
                "image": image,
                "post_id": page.get("post_id"),
                "post_link": page.get("post_link"),
                "raw": judged.get("raw"),
                "run_id": run_id,
            }
        )
        n_ok += 1
        rec_one = str(row.get("recommend_ground_truth") or "").replace("\n", " / ")
        print(
            f"{LOG} glm {i}/{len(keys)} {page.get('post_id')} pick={row['phases']['glm']['pick']} "
            f"src={row['phases']['glm']['recommend_source']} gt={rec_one[:80]!r}",
            flush=True,
        )
        if len(pending) >= max(1, flush_posts):
            _flush_glm()
        if i < len(keys):
            time.sleep(1.0)
    _flush_glm()
    glm_sum = _glm_summary_from_rows(list(existing.values()), run_id, model)
    glm_sum["processed_ok"] = n_ok
    glm_sum["processed_fail"] = n_fail
    glm_sum["recommend"] = rec_key
    upload_json_payload(bucket, sum_key, glm_sum)
    print(f"{LOG} glm done {glm_sum}", flush=True)
    return glm_sum


def _page_object_key(root: str, post_id: str, image: str) -> str:
    """Per-image fuse page key / Key page fuse từng ảnh."""
    safe = str(image).split("/")[-1]
    return f"{root}/pages/{post_id}/{safe}.json"


def _ocr_fuse_one(
    *,
    item: dict[str, str],
    bucket: str,
    root: str,
    vision_model: str,
    gpt_model: str,
    eval_model: str,
    paddle_url: str,
    batch_seq: int,
) -> dict[str, Any]:
    """Gemini∥Paddle → GPT∥ → DS → fuse one image / Một ảnh: Gemini∥Paddle → GPT∥ → DS → fuse."""
    image = str(item.get("image") or "")
    post_id = str(item.get("post_id") or "") or "unknown"
    src_key = str(item.get("src_key") or "")
    post_link = facebook_post_link(
        str(item.get("group_id") or ""),
        post_id,
        str(item.get("post_link") or ""),
    )
    caption = str(item.get("caption") or "")
    flags: list[str] = []
    try:
        raw_bytes = _read_object_bytes(bucket, src_key)
    except Exception as exc:
        page = {
            "schema_version": SCHEMA,
            "image": image,
            "post_id": post_id,
            "post_link": post_link,
            "caption": caption,
            "quote_batch": batch_seq,
            "flags": ["ocr_unreadable"],
            "error": str(exc),
            "at": utc_now_iso(),
        }
        b2 = to_task_b2_row(
            image=image,
            ground_truth="",
            side_matter="",
            gemini=[],
            post_link=post_link,
            label=caption,
        )
        return {
            "ok": False,
            "page": page,
            "b2": b2,
            "flags_row": {
                "image": image,
                "post_id": post_id,
                "post_link": post_link,
                "flags": ["ocr_unreadable"],
                "status": "",
            },
            "image_bytes": b"",
        }

    _copy_src_image(bucket, src_key, f"{root}/images/{_image_rel(image)}")
    # Overlap Gemini vision with Paddle FIFO wait / Chồng Gemini với lúc chờ hàng Paddle
    with ThreadPoolExecutor(max_workers=2) as pool:
        fut_g = pool.submit(_gemini_with_retry, raw_bytes, vision_model, 3)
        fut_p = pool.submit(_paddle_track, raw_bytes, paddle_url)
        gem = fut_g.result()
        p_blocks, p_conf, p_err, p_wait_ms, p_infer_ms = fut_p.result()

    g_boxes: list[dict[str, Any]] = gem.get("boxes") or []
    g_ink = [b for b in g_boxes if b.get("kind") == "ink_text"]
    g_conf = _clamp_conf(gem.get("confidence"), 0.0)
    if any(UI_RE.search(str(b.get("text") or "")) for b in g_boxes):
        flags.append("ui_chrome")
    if is_per_char_boxes(g_ink):
        flags.append("per_char_boxes")
    if not g_ink:
        flags.append("ocr_unreadable" if g_boxes else "not_calligraphy")
        if any(b.get("kind") == "seal" for b in g_boxes) and not g_ink:
            flags.append("seal_only")
    if g_conf <= 0.45 and g_ink:
        flags.append("ocr_suspect")
    if p_err:
        flags.append("paddle_empty")

    p_ink = [b for b in p_blocks if b.get("kind") == "ink_text"]
    if p_ink:
        p_conf = _weighted_conf(p_ink) or p_conf
    elif not p_err:
        flags.append("paddle_empty")
        p_conf = 0.0
    g_cjk_n = sum(len(compact_cjk(str(b.get("text") or ""))) for b in g_ink)
    if (p_err or not p_ink) and g_cjk_n <= 2:
        flags.append("weak_evidence")

    g_lines = cluster_boxes_to_columns(g_ink)
    p_lines = cluster_boxes_to_columns(p_ink)
    # GPT A and B in parallel / GPT A và B song song
    with ThreadPoolExecutor(max_workers=2) as pool:
        fut_a = pool.submit(_refine_lines, g_lines, gpt_model)
        fut_b = pool.submit(_refine_lines, p_lines, gpt_model) if p_lines else None
        g_ref, gpt_a_conf, f1 = fut_a.result()
        if fut_b:
            p_ref, gpt_b_conf, f2 = fut_b.result()
        else:
            p_ref, gpt_b_conf, f2 = [], 0.0, []
    flags.extend(f1)
    flags.extend(f2)
    g_eval, ds_a_conf, order_a, f3, p_eval, ds_b_conf, order_b, f4 = _eval_or_passthrough(
        g_ref, p_ref, eval_model
    )
    flags.extend(f3)
    flags.extend(f4)
    p_eval = keep_overlapping_lines(g_eval, p_eval)
    g_before, p_before = list(g_eval), list(p_eval)
    g_eval, p_eval = recover_wrapped_lines(g_eval, p_eval)
    wrap_recovered = g_eval != g_before or p_eval != p_before

    aligned = _align_line_lists(g_eval, p_eval)
    caption = str(item.get("caption") or "")
    fused = fuse_aligned_lines(aligned, caption=caption, g_cluster=g_lines)
    aligned = fused["aligned"]
    page_conf = 0.0
    if aligned:
        scored = [r for r in aligned if r.get("status") not in {"a_only", "b_only"}]
        base = scored or aligned
        page_conf = _clamp_conf(sum(float(r.get("line_conf") or 0) for r in base) / len(base))
    page_bag = bag_dice("\n".join(g_eval), "\n".join(p_eval))
    page_cer = cer("\n".join(g_eval), "\n".join(p_eval))
    cluster_bag = bag_dice("\n".join(g_lines), "\n".join(p_lines))
    cluster_cer = cer("\n".join(g_lines), "\n".join(p_lines))
    compact_bag = bag_dice("".join(g_eval), "".join(p_eval))
    flags = collect_page_flags(
        flags,
        aligned=aligned,
        fused=fused,
        wrap_recovered=wrap_recovered,
        g_eval=g_eval,
        p_eval=p_eval,
        g_boxes=g_boxes,
        p_boxes=p_blocks,
    )
    page_status = decide_page_status(
        g_eval=g_eval,
        p_eval=p_eval,
        aligned=aligned,
        page_conf=page_conf,
        page_bag=page_bag,
        page_cer=page_cer,
        cluster_bag=cluster_bag,
        cluster_cer=cluster_cer,
        flags=flags,
        p_err=p_err,
        compact_bag=compact_bag,
    )

    gt_lines, gt_track = fused["gt_lines"], fused["gt_track"]
    if "weak_evidence" in flags or "ui_chrome" in flags:
        gt_lines, gt_track = [], "none"
    gt = "\n".join(to_traditional(x) for x in gt_lines if str(x).strip())
    sm = _side_text(g_boxes + p_blocks)
    b2 = to_task_b2_row(
        image=image,
        ground_truth=gt,
        side_matter=sm,
        gemini=g_boxes,
        post_link=post_link,
        label=caption,
    )
    page = {
        "schema_version": SCHEMA,
        "image": image,
        "post_id": post_id,
        "post_link": post_link,
        "caption": caption,
        "quote_batch": batch_seq,
        "page_status": page_status,
        "flags": sorted(set(flags)),
        "boxes": {"gemini": g_boxes, "paddle": p_blocks},
        "gemini": {
            "confidence": g_conf,
            "conf_source": "model_json",
            "n_ink": len(g_ink),
            "n_cols": len(g_lines),
            "attempts": gem.get("attempts") or [],
        },
        "paddle": {
            "confidence": p_conf,
            "conf_source": "detector",
            "n_ink": len(p_ink),
            "n_cols": len(p_lines),
            "error": p_err,
            "wait_ms": p_wait_ms,
            "infer_ms": p_infer_ms,
        },
        "gpt_a": {"confidence": gpt_a_conf, "conf_source": "model_json"},
        "gpt_b": {"confidence": gpt_b_conf, "conf_source": "model_json"},
        "ds_a": {"confidence": ds_a_conf, "conf_source": "model_json", "reading_order": order_a},
        "ds_b": {"confidence": ds_b_conf, "conf_source": "model_json", "reading_order": order_b},
        "align": {
            "page_conf": page_conf,
            "conf_source": "metric",
            "cer": page_cer,
            "wer_cols": wer_lines(g_eval, p_eval),
            "ratio": fuzz_ratio("\n".join(g_eval), "\n".join(p_eval)),
            "bag": page_bag,
            "cluster_bag": cluster_bag,
            "cluster_cer": cluster_cer,
            "compact_bag": compact_bag,
            "gt_track": gt_track,
            "caption_align": fused["caption_align"],
        },
        "calligraphy_lines": aligned,
        "hitl_spans": fused["hitl_spans"],
        "locked_lines": fused["locked_lines"],
        "seals": [b for b in g_boxes + p_blocks if b.get("kind") == "seal"],
        "non_callig": [
            b for b in g_boxes + p_blocks if b.get("kind") in {"printed", "other", "margin"}
        ],
        "at": utc_now_iso(),
    }
    print(
        f"{LOG} batch={batch_seq:02d} image={image[-40:]} status={page_status} "
        f"g_conf={g_conf} p_conf={p_conf} ink={len(g_ink)}/{len(p_ink)} "
        f"cols={len(g_lines)}/{len(p_lines)} gt={gt_track} bag={page_bag} "
        f"cbag={cluster_bag} compact={compact_bag} cap={fused['caption_align']}",
        flush=True,
    )
    return {
        "ok": True,
        "page": page,
        "b2": b2,
        "flags_row": {
            "image": image,
            "post_id": post_id,
            "post_link": post_link,
            "flags": page["flags"],
            "status": page_status,
        },
        "image_bytes": raw_bytes,
    }


def _glm_row_from_fused(
    *,
    fused: dict[str, Any],
    model: str,
    run_id: str,
) -> tuple[dict[str, Any], dict[str, Any]]:
    """GLM after fuse of this image / GLM sau fuse của đúng ảnh này."""
    page = fused["page"]
    b2 = fused["b2"]
    image = str(page.get("image") or "")
    image_bytes = fused.get("image_bytes") or b""
    text_a, text_b = _page_ab_text(page)
    judged = judge_glm_page(image_bytes, text_a, text_b, model)
    row = build_glm_log_row(
        page=page,
        b2=b2,
        recommend=str(judged.get("recommend") or ""),
        pick=str(judged.get("pick") or "parse_fail"),
        source=str(judged.get("source") or "model"),
        glm_conf=float(judged.get("conf") or 0.0),
        retried=bool(judged.get("retried")),
        dropped_invented=bool(judged.get("dropped")),
        latency_ms=int(judged.get("latency_ms") or 0),
        run_id=run_id,
        model=model,
    )
    rec_one = str(row.get("recommend_ground_truth") or "").replace("\n", " / ")
    print(
        f"{LOG} glm {page.get('post_id')} pick={row['phases']['glm']['pick']} "
        f"src={row['phases']['glm']['recommend_source']} gt={rec_one[:80]!r}",
        flush=True,
    )
    raw_row = {
        "image": image,
        "post_id": page.get("post_id"),
        "post_link": page.get("post_link"),
        "raw": judged.get("raw"),
        "run_id": run_id,
    }
    return row, raw_row


def _upsert_locked(bucket: str, key: str, incoming: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """MinIO upsert by image under process lock / Upsert MinIO theo image, khóa process."""
    with _UPSERT_LOCK:
        merged = upsert_rows_by_image(read_jsonl(bucket, key), incoming)
        write_jsonl(bucket, key, merged)
        return merged


def _process_quote_batch(
    *,
    seq: int,
    items: list[dict[str, str]],
    bucket: str,
    root: str,
    vision_model: str,
    gpt_model: str,
    eval_model: str,
    paddle_url: str,
    force: bool,
    run_glm: bool,
    glm_force: bool,
    glm_model: str,
    glm_run_id: str,
    flush_posts: int,
    group_id: str = "",
) -> dict[str, Any]:
    """One quote queue: OCR+fuse then GLM overlapping the next image.

    Một queue quote: OCR+fuse rồi GLM chồng ảnh kế.
    """
    root = _assert_pilot_write_root(root)
    part_b2 = f"{root}/parts/b2/batch_{seq:02d}.jsonl"
    part_flags = f"{root}/parts/flags/batch_{seq:02d}.jsonl"
    part_glm = f"{root}/parts/glm/batch_{seq:02d}.jsonl"
    b2_key = f"{root}/task_b2.jsonl"
    flags_key = f"{root}/flags.jsonl"
    rec_key = f"{root}/glm/recommend.jsonl"
    raw_key = f"{root}/glm/runs/{glm_run_id}.jsonl"
    existing_glm = {
        str(r.get("image") or "")
        for r in read_jsonl(bucket, rec_key)
        if r.get("image")
    }
    b2_by_image = {
        str(r.get("image") or ""): r
        for r in read_jsonl(bucket, b2_key)
        if r.get("image")
    }
    batch_b2: dict[str, dict[str, Any]] = {}
    batch_flags: dict[str, dict[str, Any]] = {}
    batch_glm: dict[str, dict[str, Any]] = {}
    pending_b2: list[dict[str, Any]] = []
    pending_flags: list[dict[str, Any]] = []
    pending_glm: list[dict[str, Any]] = []
    pending_raw: list[dict[str, Any]] = []
    processed = 0
    skipped = 0
    glm_ok = 0

    def _flush_parts() -> None:
        if pending_b2:
            write_jsonl(bucket, part_b2, list(batch_b2.values()))
            _upsert_locked(bucket, b2_key, [b2_public_row(r) for r in pending_b2])
        if pending_flags:
            write_jsonl(bucket, part_flags, list(batch_flags.values()))
            _upsert_locked(bucket, flags_key, pending_flags)
        if pending_glm:
            write_jsonl(bucket, part_glm, list(batch_glm.values()))
            _upsert_locked(bucket, rec_key, pending_glm)
            with _UPSERT_LOCK:
                prev_raw = read_jsonl(bucket, raw_key)
                write_jsonl(bucket, raw_key, prev_raw + pending_raw)
        n_b2, n_glm = len(pending_b2), len(pending_glm)
        pending_b2.clear()
        pending_flags.clear()
        pending_glm.clear()
        pending_raw.clear()
        if n_b2 or n_glm:
            print(
                f"{LOG} flush batch={seq:02d} b2={n_b2} glm={n_glm} "
                f"part_b2={len(batch_b2)}",
                flush=True,
            )

    glm_pool = ThreadPoolExecutor(max_workers=1)
    prev_glm: Future | None = None
    try:
        for item in items:
            image = str(item.get("image") or "")
            post_id = str(item.get("post_id") or "") or "unknown"
            post_link = facebook_post_link(
                group_id or str(item.get("group_id") or ""),
                post_id,
                str(item.get("post_link") or ""),
            )
            item["post_link"] = post_link
            if group_id:
                item["group_id"] = group_id
            page_key = _page_object_key(root, post_id, image)
            fused: dict[str, Any] | None = None
            if not force and object_exists(bucket, page_key):
                skipped += 1
                page = _read_json_object(bucket, page_key) or {}
                page["post_link"] = str(page.get("post_link") or "") or post_link
                caption = str(item.get("caption") or page.get("caption") or "")
                page["caption"] = str(page.get("caption") or "") or caption
                if run_glm and (glm_force or image not in existing_glm):
                    try:
                        img_bytes = _read_object_bytes(
                            bucket, f"{root}/images/{_image_rel(image)}"
                        )
                    except Exception:
                        img_bytes = b""
                    prev_b2 = b2_by_image.get(image) or to_task_b2_row(
                        image=image,
                        ground_truth="",
                        side_matter="",
                        gemini=[],
                        post_link=page["post_link"],
                        label=caption,
                    )
                    if not prev_b2.get("post_link"):
                        prev_b2 = {**prev_b2, "post_link": page["post_link"]}
                    if caption and not prev_b2.get("label"):
                        prev_b2 = {**prev_b2, "label": caption}
                    fused = {
                        "page": page,
                        "b2": prev_b2,
                        "image_bytes": img_bytes,
                    }
            else:
                fused = _ocr_fuse_one(
                    item=item,
                    bucket=bucket,
                    root=root,
                    vision_model=vision_model,
                    gpt_model=gpt_model,
                    eval_model=eval_model,
                    paddle_url=paddle_url,
                    batch_seq=seq,
                )
                upload_json_payload(bucket, page_key, fused["page"])
                batch_b2[image] = b2_public_row(fused["b2"])
                b2_by_image[image] = batch_b2[image]
                batch_flags[image] = fused["flags_row"]
                pending_b2.append(batch_b2[image])
                pending_flags.append(fused["flags_row"])
                processed += 1

            need_glm = bool(
                run_glm
                and fused
                and fused.get("image_bytes")
                and (glm_force or image not in existing_glm)
            )
            # Wait GLM n-1 after OCR n so GLM n overlaps OCR n+1 /
            # Chờ GLM n-1 sau OCR n để GLM n chồng OCR n+1
            if prev_glm is not None:
                glm_row, raw_row = prev_glm.result()
                img = str(glm_row.get("image") or "")
                batch_glm[img] = glm_row
                pending_glm.append(glm_row)
                pending_raw.append(raw_row)
                existing_glm.add(img)
                glm_ok += 1
                prev_glm = None
            if need_glm:
                prev_glm = glm_pool.submit(
                    _glm_row_from_fused, fused=fused, model=glm_model, run_id=glm_run_id
                )
            if len(pending_b2) >= max(1, flush_posts):
                _flush_parts()
        if prev_glm is not None:
            glm_row, raw_row = prev_glm.result()
            img = str(glm_row.get("image") or "")
            batch_glm[img] = glm_row
            pending_glm.append(glm_row)
            pending_raw.append(raw_row)
            glm_ok += 1
        _flush_parts()
    finally:
        glm_pool.shutdown(wait=True)
    # Drop in-memory batch maps after flush / Xóa map batch trong RAM sau flush
    batch_b2.clear()
    batch_flags.clear()
    batch_glm.clear()
    return {
        "batch_seq": seq,
        "processed": processed,
        "skipped": skipped,
        "glm_ok": glm_ok,
        "n_items": len(items),
    }


def _read_page_object_safe(bucket: str, object_key: str) -> dict[str, Any] | None:
    """Read page JSON; skip empty/corrupt objects / Đọc page JSON; bỏ qua object rỗng/hỏng."""
    try:
        from common.io_storage import read_object_text

        raw = read_object_text(bucket, object_key)
        if not raw or not str(raw).strip():
            return None
        page = json.loads(raw)
        return page if isinstance(page, dict) else None
    except Exception:
        return None


def _b2_flags_from_page(page: dict[str, Any], *, caption: str = "") -> tuple[dict[str, Any], dict[str, Any]]:
    """Rebuild B2 + flags rows from a saved fuse page / Tái tạo B2+flags từ page đã fuse."""
    image = str(page.get("image") or "")
    post_id = str(page.get("post_id") or "")
    post_link = str(page.get("post_link") or "")
    cap = str(caption or page.get("caption") or "")
    stored = page.get("boxes") if isinstance(page.get("boxes"), dict) else {}
    g_boxes = list(stored.get("gemini") or [])
    p_boxes = list(stored.get("paddle") or [])
    if page.get("calligraphy_lines"):
        rec = recompute_from_page(page, caption=cap)
        gt = str(rec.get("ground_truth") or "")
        status = str(rec.get("page_status") or "")
        flags = [str(x) for x in (rec.get("flags") or page.get("flags") or [])]
    else:
        gt = ""
        status = str(page.get("page_status") or "")
        flags = [str(x) for x in (page.get("flags") or [])]
    b2 = b2_public_row(
        to_task_b2_row(
            image=image,
            ground_truth=gt,
            side_matter=_side_text(g_boxes + p_boxes),
            gemini=g_boxes,
            post_link=post_link,
            label=cap,
        )
    )
    flags_row = {
        "image": image,
        "post_id": post_id,
        "post_link": post_link,
        "flags": flags,
        "status": status,
    }
    return b2, flags_row


def backfill_gapped_sidecars(
    *,
    bucket: str,
    root: str,
    group_id: str,
    run_glm: bool = True,
    require_both: bool = True,
    flush_posts: int = FLUSH_POSTS,
    glm_model: str | None = None,
) -> dict[str, Any]:
    """Backfill task_b2 + flags + glm for pages missing sidecars after flush gaps.

    Bù task_b2 + flags + glm cho page thiếu sidecar sau lỗi flush/restart.
    """
    _ = group_id  # reserved for caption map if needed later / dự phòng map caption sau
    root = _assert_pilot_write_root(root)
    pages_prefix = f"{root}/pages/"
    b2_key = f"{root}/task_b2.jsonl"
    flags_key = f"{root}/flags.jsonl"
    rec_key = f"{root}/glm/recommend.jsonl"
    sum_key = f"{root}/glm/summary.json"
    glm_model = (glm_model or os.environ.get("FEN_LABEL_GLM_MODEL") or GLM_MODEL).strip()
    run_id = os.environ.get("FEN_LABEL_GLM_RUN_ID") or f"backfill_{time.strftime('%Y%m%d_%H%M%S')}"
    raw_key = f"{root}/glm/runs/{run_id}.jsonl"

    b2_by = {
        str(r.get("image") or ""): r for r in read_jsonl(bucket, b2_key) if r.get("image")
    }
    glm_by = {
        str(r.get("image") or ""): r for r in read_jsonl(bucket, rec_key) if r.get("image")
    }

    targets: list[tuple[str, dict[str, Any]]] = []
    skipped_bad = 0
    for key in list_objects_with_prefix(bucket, pages_prefix, suffix=".json"):
        if not str(key).startswith(pages_prefix):
            continue
        page = _read_page_object_safe(bucket, key)
        if not page:
            skipped_bad += 1
            continue
        image = str(page.get("image") or "")
        miss_b2 = image not in b2_by
        miss_glm = image not in glm_by
        if require_both:
            if not (miss_b2 and miss_glm):
                continue
        elif not (miss_b2 or miss_glm):
            continue
        targets.append((image, page))

    print(
        f"{LOG} backfill targets={len(targets)} require_both={require_both} "
        f"run_glm={run_glm} skipped_bad={skipped_bad}",
        flush=True,
    )

    pending_b2: list[dict[str, Any]] = []
    pending_flags: list[dict[str, Any]] = []
    pending_glm: list[dict[str, Any]] = []
    pending_raw: list[dict[str, Any]] = []
    n_b2 = n_glm = n_glm_fail = 0

    def _flush_backfill() -> None:
        if pending_b2:
            _upsert_locked(bucket, b2_key, pending_b2)
            pending_b2.clear()
        if pending_flags:
            _upsert_locked(bucket, flags_key, pending_flags)
            pending_flags.clear()
        if pending_glm:
            _upsert_locked(bucket, rec_key, pending_glm)
            with _UPSERT_LOCK:
                prev_raw = read_jsonl(bucket, raw_key)
                write_jsonl(bucket, raw_key, prev_raw + pending_raw)
            pending_glm.clear()
            pending_raw.clear()

    for i, (image, page) in enumerate(targets, start=1):
        miss_b2 = image not in b2_by
        miss_glm = image not in glm_by
        cap = str(page.get("caption") or "")
        if miss_b2:
            b2_row, flags_row = _b2_flags_from_page(page, caption=cap)
            pending_b2.append(b2_row)
            pending_flags.append(flags_row)
            b2_by[image] = b2_row
            n_b2 += 1
        if run_glm and miss_glm:
            text_a, text_b = _page_ab_text(page)
            if text_a or text_b:
                try:
                    image_bytes = _read_object_bytes(
                        bucket, f"{root}/images/{_image_rel(image)}"
                    )
                except Exception as exc:
                    print(f"{LOG} backfill glm miss {image} {exc}", flush=True)
                    n_glm_fail += 1
                else:
                    judged = judge_glm_page(image_bytes, text_a, text_b, glm_model)
                    row = build_glm_log_row(
                        page=page,
                        b2=b2_by.get(image),
                        recommend=str(judged.get("recommend") or ""),
                        pick=str(judged.get("pick") or "parse_fail"),
                        source=str(judged.get("source") or "model"),
                        glm_conf=float(judged.get("conf") or 0.0),
                        retried=bool(judged.get("retried")),
                        dropped_invented=bool(judged.get("dropped")),
                        latency_ms=int(judged.get("latency_ms") or 0),
                        run_id=run_id,
                        model=glm_model,
                    )
                    pending_glm.append(row)
                    pending_raw.append(
                        {
                            "image": image,
                            "post_id": page.get("post_id"),
                            "post_link": page.get("post_link"),
                            "raw": judged.get("raw"),
                            "run_id": run_id,
                        }
                    )
                    glm_by[image] = row
                    n_glm += 1
            else:
                n_glm_fail += 1
        if len(pending_b2) >= max(1, flush_posts) or len(pending_glm) >= max(1, flush_posts):
            _flush_backfill()
        if run_glm and i < len(targets):
            time.sleep(0.5)
    _flush_backfill()

    public_rows = list(b2_by.values())
    with tempfile.TemporaryDirectory(prefix="fen-b2-backfill-") as tmp:
        xlsx = f"{tmp}/task_b2.xlsx"
        _build_b2_xlsx(public_rows, xlsx)
        upload_file(get_minio_client(), bucket, f"{root}/task_b2.xlsx", xlsx)
    if run_glm and glm_by:
        glm_sum = _glm_summary_from_rows(list(glm_by.values()), run_id, glm_model)
        glm_sum["backfill_ok"] = n_glm
        glm_sum["backfill_fail"] = n_glm_fail
        upload_json_payload(bucket, sum_key, glm_sum)

    summary = {
        "schema_version": SCHEMA,
        "backfill": True,
        "require_both": require_both,
        "targets": len(targets),
        "n_b2_upserted": n_b2,
        "n_glm_upserted": n_glm,
        "n_glm_fail": n_glm_fail,
        "skipped_bad_pages": skipped_bad,
        "n_b2_total": len(b2_by),
        "n_glm_total": len(glm_by),
        "task_b2": b2_key,
        "glm_recommend": rec_key,
        "updated_at": utc_now_iso(),
    }
    print(f"{LOG} backfill done {summary}", flush=True)
    return summary


def merge_export_prefix(root: str) -> str:
    """Export folder for merged B2 under the pilot prefix / Thư mục export B2 gộp."""
    root = _assert_pilot_write_root(root)
    return f"{root}/export"


def _norm_b2_image_path(image: str) -> str:
    """Normalize Task B2 image path / Chuẩn hóa đường dẫn ảnh B2."""
    img = str(image or "").strip()
    if not img:
        return ""
    return img if img.startswith("/") else f"/{img.lstrip('/')}"


def load_quote_universe(bucket: str, root: str) -> dict[str, dict[str, Any]]:
    """Map image → quote queue metadata for all 12 batches.

    Ánh xạ image → metadata queue quote cho cả 12 batch.
    """
    root = _assert_pilot_write_root(root)
    out: dict[str, dict[str, Any]] = {}
    for seq in range(1, QUOTE_BATCH_COUNT + 1):
        for row in read_jsonl(bucket, quote_queue_key(root, seq)):
            image = _norm_b2_image_path(str(row.get("image") or ""))
            if not image:
                continue
            out[image] = {
                "quote_batch": seq,
                "caption": str(row.get("caption") or ""),
                "post_link": str(row.get("post_link") or ""),
                "post_id": str(row.get("post_id") or ""),
            }
    return out


def _legacy_row_to_b2(legacy: dict[str, Any], meta: dict[str, Any]) -> dict[str, Any]:
    """Map legacy ocr_result row to public B2 / Map dòng ocr_result cũ sang B2 public."""
    gemini = legacy.get("gemini") if isinstance(legacy.get("gemini"), list) else []
    return b2_public_row(
        to_task_b2_row(
            image=_norm_b2_image_path(str(legacy.get("image") or "")),
            ground_truth=str(legacy.get("ground_truth") or ""),
            side_matter=str(legacy.get("side_matter") or ""),
            gemini=gemini,
            post_link=str(meta.get("post_link") or legacy.get("post_link") or ""),
            label=str(meta.get("caption") or legacy.get("label") or ""),
        )
    )


def _fuse_row_to_b2(fuse: dict[str, Any], meta: dict[str, Any]) -> dict[str, Any]:
    """Normalize fuse task_b2 row with queue caption/link / Chuẩn hóa fuse B2 với caption queue."""
    row = b2_public_row(fuse)
    if not str(row.get("label") or "").strip():
        row["label"] = str(meta.get("caption") or "")
    if not str(row.get("post_link") or "").strip():
        row["post_link"] = str(meta.get("post_link") or "")
    return row


def resolve_merge_b2_row(
    *,
    image: str,
    meta: dict[str, Any],
    fuse_row: dict[str, Any] | None,
    legacy_row: dict[str, Any] | None,
    prefer: str = "fuse_first",
    min_bag_warn: float = 0.70,
) -> tuple[dict[str, Any], dict[str, Any]]:
    """Pick merged B2 row + manifest entry for one quote image.

    Chọn dòng B2 gộp + manifest cho một ảnh quote.
    """
    fuse_gt = str((fuse_row or {}).get("ground_truth") or "").strip()
    legacy_gt = str((legacy_row or {}).get("ground_truth") or "").strip()
    both_present = bool(fuse_gt and legacy_gt)
    bag = bag_dice(legacy_gt, fuse_gt) if both_present else None
    diverge = bool(both_present and bag is not None and bag < float(min_bag_warn))

    # fuse_first: non-empty fuse GT wins; empty fuse falls back to legacy /
    # fuse_first: fuse GT khác rỗng thắng; fuse rỗng thì legacy
    if prefer == "fuse_first":
        if fuse_gt:
            gt_source = "fuse"
            chosen = _fuse_row_to_b2(fuse_row or {}, meta)
        elif legacy_gt:
            gt_source = "legacy_ocr"
            chosen = _legacy_row_to_b2(legacy_row or {}, meta)
        else:
            gt_source = "pending"
            chosen = b2_public_row(
                to_task_b2_row(
                    image=_norm_b2_image_path(image),
                    ground_truth="",
                    side_matter="",
                    gemini=[],
                    post_link=str(meta.get("post_link") or ""),
                    label=str(meta.get("caption") or ""),
                )
            )
    elif prefer == "legacy_first":
        if legacy_gt:
            gt_source = "legacy_ocr"
            chosen = _legacy_row_to_b2(legacy_row or {}, meta)
        elif fuse_gt:
            gt_source = "fuse"
            chosen = _fuse_row_to_b2(fuse_row or {}, meta)
        else:
            gt_source = "pending"
            chosen = b2_public_row(
                to_task_b2_row(
                    image=_norm_b2_image_path(image),
                    ground_truth="",
                    side_matter="",
                    gemini=[],
                    post_link=str(meta.get("post_link") or ""),
                    label=str(meta.get("caption") or ""),
                )
            )
    else:
        raise ValueError(f"unsupported merge prefer={prefer!r}")

    manifest = {
        "image": _norm_b2_image_path(image),
        "post_id": str(meta.get("post_id") or ""),
        "quote_batch": int(meta.get("quote_batch") or 0),
        "post_link": str(meta.get("post_link") or ""),
        "label_caption": str(meta.get("caption") or ""),
        "gt_source": gt_source,
        "ground_truth": str(chosen.get("ground_truth") or ""),
        "fuse_gt": fuse_gt,
        "legacy_gt": legacy_gt,
        "both_present": both_present,
        "bag_fuse_vs_legacy": bag,
        "diverge_flag": diverge,
    }
    return chosen, manifest


def merge_task_b2_sources(
    *,
    bucket: str,
    root: str,
    source_prefix: str,
    group_id: str,
    prefer: str = "fuse_first",
    min_bag_warn: float = 0.70,
    dry_run: bool = False,
) -> dict[str, Any]:
    """Merge fuse task_b2 with legacy ocr_result for all 12 quote batches.

    Gộp task_b2 fuse với ocr_result cũ cho toàn bộ 12 batch quote.
    """
    root = _assert_pilot_write_root(root)
    gid = str(group_id or "").strip()
    export_root = merge_export_prefix(root)
    universe = load_quote_universe(bucket, root)
    fuse_by = {
        _norm_b2_image_path(str(r.get("image") or "")): r
        for r in read_jsonl(bucket, f"{root}/task_b2.jsonl")
        if _norm_b2_image_path(str(r.get("image") or ""))
    }
    legacy_by = {
        _norm_b2_image_path(str(r.get("image") or "")): r
        for r in read_jsonl(bucket, _ocr_output_key(source_prefix, gid))
        if _norm_b2_image_path(str(r.get("image") or ""))
    }

    merged_rows: list[dict[str, Any]] = []
    manifest_rows: list[dict[str, Any]] = []
    source_counts: Counter[str] = Counter()
    diverge_count = 0
    for image in sorted(
        universe,
        key=lambda img: (int(universe[img].get("quote_batch") or 0), img),
    ):
        meta = universe[image]
        chosen, manifest = resolve_merge_b2_row(
            image=image,
            meta=meta,
            fuse_row=fuse_by.get(image),
            legacy_row=legacy_by.get(image),
            prefer=prefer,
            min_bag_warn=min_bag_warn,
        )
        merged_rows.append(chosen)
        manifest_rows.append(manifest)
        source_counts[str(manifest.get("gt_source") or "pending")] += 1
        if manifest.get("diverge_flag"):
            diverge_count += 1

    out_jsonl = f"{export_root}/task_b2_merged.jsonl"
    out_xlsx = f"{export_root}/task_b2_merged.xlsx"
    out_manifest = f"{export_root}/merge_manifest.jsonl"
    out_summary = f"{export_root}/merge_summary.json"
    summary = {
        "schema_version": SCHEMA,
        "merge": True,
        "prefer": prefer,
        "min_bag_warn": float(min_bag_warn),
        "dry_run": bool(dry_run),
        "n_quote_universe": len(universe),
        "n_merged": len(merged_rows),
        "n_fuse_rows": len(fuse_by),
        "n_legacy_rows": len(legacy_by),
        "gt_source_counts": dict(source_counts),
        "n_diverge_flag": diverge_count,
        "task_b2_merged": out_jsonl,
        "task_b2_merged_xlsx": out_xlsx,
        "merge_manifest": out_manifest,
        "updated_at": utc_now_iso(),
    }
    if not dry_run:
        write_jsonl(bucket, out_jsonl, merged_rows)
        write_jsonl(bucket, out_manifest, manifest_rows)
        with tempfile.TemporaryDirectory(prefix="fen-b2-merge-") as tmp:
            xlsx = f"{tmp}/task_b2_merged.xlsx"
            _build_b2_xlsx(merged_rows, xlsx)
            upload_file(get_minio_client(), bucket, out_xlsx, xlsx)
        upload_json_payload(bucket, out_summary, summary)
    print(f"{LOG} merge done {summary}", flush=True)
    return summary


def run_label_dual_merge(
    *,
    group_id: str | None = None,
    prefer: str = "fuse_first",
    min_bag_warn: float = 0.70,
    dry_run: bool = False,
) -> dict[str, Any]:
    """Scheduled merge of fuse B2 + legacy OCR for quote universe.

    Gộp định kỳ B2 fuse + OCR cũ cho universe quote (read-only legacy).
    """
    gid = (group_id or os.environ.get("FEN_GROUP_ID") or "322453387859386").strip()
    bucket, source_prefix = ensure_raw_bucket()
    root = label_dual_root(source_prefix, gid)
    if os.environ.get("FEN_LABEL_MERGE_PREFER"):
        prefer = os.environ.get("FEN_LABEL_MERGE_PREFER", prefer).strip() or prefer
    if os.environ.get("FEN_LABEL_MERGE_MIN_BAG_WARN"):
        try:
            min_bag_warn = float(os.environ.get("FEN_LABEL_MERGE_MIN_BAG_WARN", str(min_bag_warn)))
        except ValueError:
            pass
    if os.environ.get("FEN_LABEL_MERGE_DRY_RUN") is not None:
        dry_run = _bool_env("FEN_LABEL_MERGE_DRY_RUN", dry_run)
    return merge_task_b2_sources(
        bucket=bucket,
        root=root,
        source_prefix=source_prefix,
        group_id=gid,
        prefer=prefer,
        min_bag_warn=min_bag_warn,
        dry_run=dry_run,
    )


def run_label_dual_backfill(
    *,
    group_id: str | None = None,
    flush_posts: int = FLUSH_POSTS,
    run_glm: bool = True,
    require_both: bool = True,
) -> dict[str, Any]:
    """One-shot sidecar backfill for pages missing B2/GLM under label_dual_pilot.

    Bù sidecar B2/GLM cho page thiếu trong label_dual_pilot (không OCR lại).
    """
    gid = (group_id or os.environ.get("FEN_GROUP_ID") or "322453387859386").strip()
    bucket, _source_prefix = ensure_raw_bucket()
    root = label_dual_root(_source_prefix, gid)
    cfg = load_config()
    glm_model = (os.environ.get("FEN_LABEL_GLM_MODEL") or GLM_MODEL).strip()
    if os.environ.get("FEN_LABEL_BACKFILL_REQUIRE_BOTH") is not None:
        require_both = _bool_env("FEN_LABEL_BACKFILL_REQUIRE_BOTH", require_both)
    if os.environ.get("FEN_LABEL_GLM") is not None:
        run_glm = _bool_env("FEN_LABEL_GLM", run_glm)
    return backfill_gapped_sidecars(
        bucket=bucket,
        root=root,
        group_id=gid,
        run_glm=run_glm,
        require_both=require_both,
        flush_posts=flush_posts,
        glm_model=glm_model,
    )


def run_label_dual(
    *,
    group_id: str | None = None,
    limit: int = 10,
    flush_posts: int = FLUSH_POSTS,
    force: bool = False,
    replay: bool = False,
    glm: bool | None = None,
    batch_seq: int | None = None,
    workers: int | None = None,
    prepare_queues: bool | None = None,
) -> dict[str, Any]:
    """Quote corpus: 12 queues, per-image Gemini∥Paddle → fuse → GLM overlap.

    Corpus quote: 12 queue, từng ảnh Gemini∥Paddle → fuse → GLM chồng.
    """
    gid = (group_id or os.environ.get("FEN_GROUP_ID") or "322453387859386").strip()
    bucket, source_prefix = ensure_raw_bucket()
    root = _assert_pilot_write_root(label_dual_root(source_prefix, gid))
    run_glm = _bool_env("FEN_LABEL_GLM", True) if glm is None else glm
    glm_force = _bool_env("FEN_LABEL_GLM_FORCE", False)
    seq = _int_env("FEN_LABEL_BATCH_SEQ", 0) if batch_seq is None else int(batch_seq)
    n_workers = (
        _int_env("FEN_LABEL_WORKERS", QUOTE_BATCH_COUNT) if workers is None else int(workers)
    )
    # Cap at 12: one pod, two GPUs; more threads only queue on Paddle /
    # Trần 12: một pod, hai GPU; thêm thread chỉ xếp hàng Paddle
    n_workers = max(1, min(QUOTE_BATCH_COUNT, n_workers))
    do_prepare = (
        _bool_env("FEN_LABEL_PREPARE_QUEUES", True)
        if prepare_queues is None
        else bool(prepare_queues)
    )
    cfg = load_config()
    vision_model = os.environ.get("FEN_LABEL_VISION_MODEL", "").strip() or get_value(
        cfg, "final_exam_nlp", "label_vision_model", fallback=VISION_MODEL
    )
    gpt_model = os.environ.get("FEN_LABEL_GPT_MODEL", "").strip() or get_value(
        cfg, "final_exam_nlp", "label_gpt_model", fallback=GPT_MODEL
    )
    eval_model = _resolve_eval_model(cfg)
    paddle_url = os.environ.get("FEN_PADDLE_OCR_URL", "").strip() or get_value(
        cfg, "final_exam_nlp", "paddle_ocr_url", fallback=PADDLE_URL
    )
    glm_model = (os.environ.get("FEN_LABEL_GLM_MODEL") or GLM_MODEL).strip()
    glm_run_id = os.environ.get("FEN_LABEL_GLM_RUN_ID") or f"glm_rec_{time.strftime('%Y%m%d_%H%M%S')}"

    def _attach_glm_pointer(summary: dict[str, Any], glm_sum: dict[str, Any] | None = None) -> dict[str, Any]:
        summary = dict(summary)
        summary["glm"] = {
            "schema_version": GLM_LOG_SCHEMA,
            "recommend": f"{root}/glm/recommend.jsonl",
            "summary": f"{root}/glm/summary.json",
            "last_run_id": (glm_sum or {}).get("run_id") or glm_run_id,
            "n": (glm_sum or {}).get("n"),
        }
        return summary

    if replay:
        summary = replay_existing_pilot_pages(
            bucket=bucket,
            root=root,
            source_prefix=source_prefix,
            group_id=gid,
        )
        if run_glm:
            glm_sum = run_glm_recommend_pass(
                bucket=bucket,
                root=root,
                flush_posts=flush_posts,
                force=glm_force,
            )
            summary = _attach_glm_pointer(summary, glm_sum)
            upload_json_payload(bucket, f"{root}/summary.json", summary)
        return summary

    queue_meta: dict[str, Any] = {}
    if do_prepare:
        queue_meta = prepare_quote_queues(
            bucket=bucket,
            root=root,
            source_prefix=source_prefix,
            group_id=gid,
            n_batches=QUOTE_BATCH_COUNT,
            force=_bool_env("FEN_LABEL_PREPARE_FORCE", False),
        )

    seqs = [seq] if seq > 0 else list(range(1, QUOTE_BATCH_COUNT + 1))
    priority_only = _bool_env("FEN_LABEL_PRIORITY_ONLY", False)
    run_priority = _bool_env("FEN_LABEL_PRIORITY_LONGLINE", True)
    # Target = max pending images this run; 0 = unlimited /
    # Target = số ảnh pending tối đa lần này; 0 = không giới hạn
    target = _int_env("FEN_LABEL_TARGET", 0)
    if target <= 0 and int(limit) > 0:
        target = int(limit)
    if (
        target <= 0
        and not priority_only
        and not force
        and not _bool_env("FEN_LABEL_UNLIMITED", False)
    ):
        # Safe default chunk for rollover (skip-done) / Chunk mặc định an toàn khi rollover
        target = DEFAULT_TARGET_CHUNK

    priority_items: list[dict[str, str]] = []
    priority_meta: dict[str, int] = {}
    if run_priority and (priority_only or seq == 0 or seq == PRIORITY_LONGLINE_SEQ):
        raw_priority = [
            r
            for r in read_jsonl(bucket, priority_longline_queue_key(root))
            if str(r.get("image") or "").strip()
        ]
        # Skip images already rewritten by priority shards; still force remaining /
        # Skip ảnh đã ghi lại bởi shard priority; ảnh còn lại vẫn force re-OCR
        pri_jobs, priority_meta = select_pending_chunk(
            bucket=bucket,
            root=root,
            queue_rows_by_seq=[(PRIORITY_LONGLINE_SEQ, raw_priority)],
            target=target,
            force=False,
            is_done=_item_priority_reocr_done,
        )
        if pri_jobs:
            priority_items = list(pri_jobs[0][1])

    jobs: list[tuple[int, list[dict[str, str]]]] = []
    chunk_meta: dict[str, int] = {}
    if not priority_only:
        queue_rows: list[tuple[int, list[dict[str, Any]]]] = []
        for s in seqs:
            if s == PRIORITY_LONGLINE_SEQ:
                continue
            queue_rows.append((s, read_jsonl(bucket, quote_queue_key(root, s))))
        jobs, chunk_meta = select_pending_chunk(
            bucket=bucket,
            root=root,
            queue_rows_by_seq=queue_rows,
            target=target,
            force=force,
        )
        # Keep 12-way if skip-done left fewer queues than workers /
        # Giữ 12-way nếu skip-done còn ít queue hơn worker
        jobs = expand_jobs_for_workers(jobs, n_workers)
        # Free queue lists not selected / Giải phóng list queue không chọn
        del queue_rows

    n_queued = sum(len(r) for _s, r in jobs)
    print(
        f"{LOG} quote jobs={len(jobs)} queued={n_queued} "
        f"priority_longline={len(priority_items)} "
        f"priority_only={priority_only} target={target} limit={limit} "
        f"chunk={chunk_meta} priority_chunk={priority_meta} "
        f"workers={n_workers} seqs={seqs} vision={vision_model} "
        f"eval={eval_model or 'disabled'} "
        f"api_keys={len(collect_api_keys())} dest={bucket}/{root}",
        flush=True,
    )

    def _run_seq(pair: tuple[int, list[dict[str, str]]], *, force_batch: bool) -> dict[str, Any]:
        s, rows = pair
        return _process_quote_batch(
            seq=s,
            items=rows,
            bucket=bucket,
            root=root,
            vision_model=vision_model,
            gpt_model=gpt_model,
            eval_model=eval_model,
            paddle_url=paddle_url,
            force=force_batch,
            run_glm=run_glm,
            glm_force=glm_force or force_batch,
            glm_model=glm_model,
            glm_run_id=glm_run_id,
            flush_posts=flush_posts,
            group_id=gid,
        )

    batch_results: list[dict[str, Any]] = []
    n_pri_workers = 1
    # Long-line anomalies first, always force re-OCR+fuse /
    # Ưu tiên GT 1 dòng dài, luôn force re-OCR+fuse
    if priority_items:
        pri_jobs = shard_items_for_workers(priority_items, n_workers)
        n_pri_workers = max(1, len(pri_jobs))
        print(
            f"{LOG} priority_longline start n={len(priority_items)} "
            f"shards={n_pri_workers} workers={n_workers} force=true",
            flush=True,
        )
        if n_pri_workers > 1:
            with ThreadPoolExecutor(max_workers=n_pri_workers) as pool:
                batch_results.extend(
                    list(pool.map(lambda j: _run_seq(j, force_batch=True), pri_jobs))
                )
        else:
            batch_results.append(_run_seq(pri_jobs[0], force_batch=True))
        priority_items.clear()

    use_pool = n_workers > 1 and len(jobs) > 1
    if use_pool:
        with ThreadPoolExecutor(max_workers=min(n_workers, len(jobs))) as pool:
            batch_results.extend(
                list(pool.map(lambda j: _run_seq(j, force_batch=force), jobs))
            )
    else:
        batch_results.extend([_run_seq(j, force_batch=force) for j in jobs])
    jobs.clear()

    final_b2 = read_jsonl(bucket, f"{root}/task_b2.jsonl")
    merged_flags = read_jsonl(bucket, f"{root}/flags.jsonl")
    n_silver = sum(1 for r in merged_flags if r.get("status") == "silver")
    if final_b2:
        with tempfile.TemporaryDirectory(prefix="fen-b2-") as tmp:
            xlsx = f"{tmp}/task_b2.xlsx"
            _build_b2_xlsx([b2_public_row(r) for r in final_b2], xlsx)
            upload_file(get_minio_client(), bucket, f"{root}/task_b2.xlsx", xlsx)
        # Pure Task B2 (no side_matter) for local submit / B2 thuần nộp local
        local_paths = write_b2_submit_local(group_id=gid, rows=final_b2)
    else:
        local_paths = {}
    glm_rows = read_jsonl(bucket, f"{root}/glm/recommend.jsonl")
    glm_sum = _glm_summary_from_rows(glm_rows, glm_run_id, glm_model) if glm_rows else {}
    if glm_sum:
        upload_json_payload(bucket, f"{root}/glm/summary.json", glm_sum)
    processed = sum(int(r.get("processed") or 0) for r in batch_results)
    skipped = sum(int(r.get("skipped") or 0) for r in batch_results)
    more_pending = bool(
        (priority_only and priority_meta.get("chunk_full"))
        or ((not priority_only) and chunk_meta.get("chunk_full"))
    )
    summary: dict[str, Any] = {
        "schema_version": SCHEMA,
        "n_images": len(final_b2),
        "processed_this_run": processed,
        "skipped_this_run": skipped,
        "n_silver": n_silver,
        "n_flagged": max(len(merged_flags) - n_silver, 0),
        "n_silver_this_run": sum(1 for r in merged_flags if r.get("status") == "silver"),
        "force": force,
        "target": target,
        "chunk": chunk_meta,
        "priority_chunk": priority_meta,
        "more_pending": more_pending,
        "batch_seq": seq,
        "workers": max(n_pri_workers, n_workers if use_pool else 1),
        "batch_results": batch_results,
        "queue": queue_meta,
        "task_b2": f"{root}/task_b2.jsonl",
        "task_b2_xlsx": f"{root}/task_b2.xlsx",
        "local_submit": local_paths,
        "images_prefix": f"{root}/images/",
        "root": root,
        "updated_at": utc_now_iso(),
    }
    if run_glm:
        summary = _attach_glm_pointer(summary, glm_sum)
    ck_out = {
        "schema_version": SCHEMA,
        "done_images": sorted(str(r.get("image") or "") for r in final_b2 if r.get("image")),
        "n_b2": len(final_b2),
        "last_flush_at": utc_now_iso(),
        "batch_seq": seq,
        "target": target,
        "more_pending": more_pending,
        "chunk": chunk_meta,
        "priority_chunk": priority_meta,
    }
    upload_json_payload(bucket, f"{root}/checkpoint.json", ck_out)
    upload_json_payload(bucket, f"{root}/summary.json", summary)
    print(f"{LOG} done {summary}", flush=True)
    return summary


def main() -> None:
    run_label_dual(
        group_id=os.environ.get("FEN_GROUP_ID") or "322453387859386",
        limit=_int_env("FEN_LABEL_LIMIT", 0),
        flush_posts=_int_env("FEN_LABEL_FLUSH_POSTS", FLUSH_POSTS),
        force=_bool_env("FEN_LABEL_FORCE", False),
        replay=_bool_env("FEN_LABEL_REPLAY", False),
        glm=_bool_env("FEN_LABEL_GLM", True),
        batch_seq=_int_env("FEN_LABEL_BATCH_SEQ", 0),
        workers=_int_env("FEN_LABEL_WORKERS", 12),
        prepare_queues=_bool_env("FEN_LABEL_PREPARE_QUEUES", True),
    )


if __name__ == "__main__":
    main()
